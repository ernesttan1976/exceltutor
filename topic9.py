# create_dates_time_workbook.py
# Builds an Excel practice file for N Level: Dates & Time
# Sheets: Instructions, Data, Tasks, Hints, Answers, Checklist, Lookup

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

# ---------- Helpers ----------
thin = Side(style="thin", color="CCCCCC")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_col_width(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_border(ws, cell_range):
    for row in ws[cell_range]:
        for c in row:
            c.border = border_thin


# ---------- Workbook ----------
wb = Workbook()

# 1) Instructions
wsI = wb.active
wsI.title = "Instructions"
wsI["A1"] = "Excel Practice: Date & Time (N Level)"
wsI["A1"].font = Font(size=14, bold=True)
wsI["A3"] = "What you’ll practice"
wsI["A4"] = "- TODAY() and NOW()"
wsI["A5"] = "- DAY(), MONTH(), YEAR()"
wsI["A6"] = "- Date formatting and simple calculations"
wsI["A8"] = "How to use this workbook"
wsI["A9"] = "1) Read Tasks sheet and follow the steps."
wsI["A10"] = "2) Use Data sheet for input and formulas."
wsI["A11"] = "3) Check Hints if you’re stuck."
wsI["A12"] = "4) Compare with Answers when done."
wsI["A14"] = (
    "Tip: If you see ##### widen the column. Right-click column header → Column Width."
)
set_col_width(wsI, {"A": 95})

# 2) Data
wsD = wb.create_sheet("Data")
headers = [
    "SampleDate",
    "Event",
    "Person",
    "DueInDays",
    "DueDate",
    "Day",
    "Month",
    "Year",
    "Today",
    "Now",
]
wsD.append(headers)

# Sample rows (spread across months)
rows = [
    ("2025-01-15", "Orientation", "Amir", 10),
    ("2025-02-03", "Lab Booking", "Bella", 7),
    ("2025-03-22", "CCA Signup", "Chen", 5),
    ("2025-04-09", "Sports Day", "Dinesh", 3),
    ("2025-05-30", "Mid-Year Exam", "Ella", 14),
    ("2025-06-18", "Camp", "Farah", 9),
    ("2025-07-05", "Project Milestone", "Gopal", 12),
    ("2025-08-12", "Parent Meeting", "Hana", 4),
    ("2025-09-21", "School Fair", "Ivan", 6),
    ("2025-10-02", "Submission", "Jia", 2),
    ("2025-11-14", "Open House", "Kai", 8),
    ("2025-12-08", "Results", "Lena", 1),
    ("2025-03-01", "Library Audit", "Maya", 11),
    ("2025-04-27", "Workshop", "Noah", 13),
    ("2025-05-06", "Assembly", "Omar", 3),
    ("2025-06-25", "ECA", "Priya", 7),
    ("2025-07-19", "Showcase", "Qin", 10),
    ("2025-08-30", "Briefing", "Ravi", 5),
    ("2025-09-10", "Competition", "Sara", 9),
    ("2025-10-29", "Cleanup", "Tariq", 4),
]
for r in rows:
    wsD.append(list(r))

# Formulas and formats
last_row = wsD.max_row
for r in range(2, last_row + 1):
    # DueDate = SampleDate + DueInDays
    wsD[f"E{r}"] = f"=A{r}+D{r}"
    # Day / Month / Year
    wsD[f"F{r}"] = f"=DAY(A{r})"
    wsD[f"G{r}"] = f"=MONTH(A{r})"
    wsD[f"H{r}"] = f"=YEAR(A{r})"
    # Today / Now
    wsD[f"I{r}"] = "=TODAY()"
    wsD[f"J{r}"] = "=NOW()"
    # Formats
    for c in ["A", "E", "I"]:
        wsD[f"{c}{r}"].number_format = "DD-MMM-YYYY"
    wsD[f"J{r}"].number_format = "DD-MMM-YYYY HH:MM"

# Style header and columns
style_header(wsD, 1)
set_col_width(
    wsD,
    {
        "A": 14,
        "B": 18,
        "C": 14,
        "D": 10,
        "E": 14,
        "F": 7,
        "G": 8,
        "H": 8,
        "I": 14,
        "J": 19,
    },
)
apply_border(wsD, f"A1:J{last_row}")

# Summary by month (K:L:M) + chart
wsD["K1"], wsD["L1"], wsD["M1"] = "MonthNum", "Month", "Count"
for i in range(2, 14):  # rows 2..13 for months 1..12
    wsD[f"K{i}"] = i - 1
    wsD[f"L{i}"] = f"=VLOOKUP(K{i},Lookup!$A$2:$B$13,2,FALSE)"
    wsD[f"M{i}"] = f"=COUNTIF($G$2:$G${last_row},K{i})"
style_header(wsD, 1)
apply_border(wsD, "K1:M13")

chart = BarChart()
chart.title = "Events by Month"
data_ref = Reference(wsD, min_col=13, min_row=1, max_row=13, max_col=13)  # M1:M13
cats_ref = Reference(wsD, min_col=12, min_row=2, max_row=13)  # L2:L13
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.y_axis.title = "Count"
chart.x_axis.title = "Month"
wsD.add_chart(chart, "O2")

# 3) Tasks
wsT = wb.create_sheet("Tasks")
wsT["A1"] = "Practice Tasks: Date & Time"
wsT["A1"].font = Font(size=13, bold=True)
tasks = [
    (
        "Starter",
        "In B2 enter =TODAY(). In C2 enter =NOW(). Format them as date and date+time.",
    ),
    ("Starter", "Given A5 has a date, extract Day in B5, Month in C5, Year in D5."),
    (
        "Core",
        "In E2:E21, DueDate is SampleDate + DueInDays. Confirm formulas already work.",
    ),
    (
        "Core",
        "Create a readable format: select A2:A21 and E2:E21 → format as DD-MMM-YYYY.",
    ),
    (
        "Core",
        "Use MONTH numbers in G2:G21 to summarise counts by month (see table in K:M).",
    ),
    (
        "Stretch",
        "Birthday age: If A10 has 01/01/2000, calculate age this year: =YEAR(TODAY())-YEAR(A10).",
    ),
    (
        "Stretch",
        "Use VLOOKUP to convert month number (K2:K13) to month name from Lookup sheet.",
    ),
    (
        "Stretch",
        "Filter Data to show only rows for a chosen Month (dropdown in B2 below).",
    ),
]
wsT.append(["Level", "Task"])
for lvl, txt in tasks:
    wsT.append([lvl, txt])
style_header(wsT, 1)
set_col_width(wsT, {"A": 12, "B": 95})
apply_border(wsT, f"A1:B{wsT.max_row}")

# Add a small interactive area for filter selection
wsT["B2"] = "Choose a Month:"
wsT["B3"] = ""  # user will choose via dropdown
dv = DataValidation(type="list", formula1="=Lookup!$B$2:$B$13", allow_blank=True)
wsT.add_data_validation(dv)
dv.add(wsT["B3"])
wsT["B5"] = "Tip: Use Data → Filter on the Data sheet and filter by the chosen month."

# 4) Hints
wsH = wb.create_sheet("Hints")
wsH["A1"] = "Hints"
wsH["A1"].font = Font(size=13, bold=True)
hints = [
    "TODAY() returns the current date; NOW() returns date + time.",
    "Extract parts: =DAY(A2), =MONTH(A2), =YEAR(A2).",
    "Due date: =A2 + D2 if D2 is days.",
    "Format dates: Ctrl+1 (Mac: Cmd+1) → Number → Date.",
    "Month name from number: =VLOOKUP(K2, Lookup!$A$2:$B$13, 2, FALSE).",
    "Count rows in a month: =COUNTIF($G$2:$G$21, K2).",
]
wsH.append(["Tip"])
for t in hints:
    wsH.append([t])
style_header(wsH, 1)
set_col_width(wsH, {"A": 95})
apply_border(wsH, f"A1:A{wsH.max_row}")

# 5) Answers
wsA = wb.create_sheet("Answers")
wsA["A1"] = "Suggested Answers (formulas)"
wsA["A1"].font = Font(size=13, bold=True)
answers = [
    ("B2 (TODAY)", "=TODAY()"),
    ("C2 (NOW)", "=NOW()"),
    ("B5 (DAY of A5)", "=DAY(Data!A5)"),
    ("C5 (MONTH of A5)", "=MONTH(Data!A5)"),
    ("D5 (YEAR of A5)", "=YEAR(Data!A5)"),
    ("E2 (DueDate)", "=Data!A2+Data!D2"),
    ("K2:K13 (Month numbers)", "1..12"),
    ("L2 (Month name)", "=VLOOKUP(Data!K2, Lookup!$A$2:$B$13, 2, FALSE)"),
    ("M2 (Count for month in K2)", "=COUNTIF(Data!$G$2:$G$21, Data!K2)"),
    ("Age this year", "=YEAR(TODAY()) - YEAR(A10)"),
]
wsA.append(["Cell / Range", "Formula"])
for label, f in answers:
    wsA.append([label, f])
style_header(wsA, 1)
set_col_width(wsA, {"A": 28, "B": 80})
apply_border(wsA, f"A1:B{wsA.max_row}")

# 6) Checklist
wsC = wb.create_sheet("Checklist")
wsC["A1"] = "Self-Check"
wsC["A1"].font = Font(size=13, bold=True)
wsC.append(["Item", "Done (Yes/No)"])
check_items = [
    "I used TODAY() and NOW().",
    "I extracted DAY/MONTH/YEAR correctly.",
    "I formatted dates as DD-MMM-YYYY.",
    "I computed DueDate = SampleDate + DueInDays.",
    "I created/understood the month summary and chart.",
]
for item in check_items:
    wsC.append([item, ""])
style_header(wsC, 1)
apply_border(wsC, f"A1:B{wsC.max_row}")
set_col_width(wsC, {"A": 60, "B": 16})
# Yes/No dropdown
dv2 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
wsC.add_data_validation(dv2)
for r in range(2, wsC.max_row + 1):
    dv2.add(wsC[f"B{r}"])

# 7) Lookup
wsL = wb.create_sheet("Lookup")
wsL.append(["MonthNum", "MonthName"])
months = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]
for i, m in enumerate(months, start=1):
    wsL.append([i, m])
style_header(wsL, 1)
apply_border(wsL, "A1:B13")
set_col_width(wsL, {"A": 10, "B": 12})

# Freeze panes & aesthetics
wsD.freeze_panes = "A2"
wsT.freeze_panes = "A2"
wsH.freeze_panes = "A2"
wsA.freeze_panes = "A2"

# Save
filename = "dates_time_practice.xlsx"
wb.save(filename)
print(f"Workbook created: {filename}")
