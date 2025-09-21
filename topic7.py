# make_lookup_workbook.py
# Creates: lookup_practice.xlsx
# Sheets: Instructions, Data, Tasks, Hints, Answers, Checklist, Lookup

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from datetime import datetime

wb = Workbook()

# Helper styles
title_font = Font(bold=True, size=14)
header_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
wrap = Alignment(wrap_text=True)
thin = Side(style="thin", color="CCCCCC")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
fill_header = PatternFill("solid", fgColor="F2F2F2")

# ---------- Sheet: Instructions ----------
ws = wb.active
ws.title = "Instructions"
ws["A1"] = "Excel Lookup Functions — Starter Workbook"
ws["A1"].font = title_font
ws["A3"] = (
    "Goal: Practice using VLOOKUP (and XLOOKUP if available) to fetch a student's Name and Grade by StudentID.\n"
    "What’s inside:\n"
    "• Data: Student list with IDs, Names, Subject, Grade (as a formatted Table)\n"
    "• Lookup: A dropdown to pick StudentID + formulas for VLOOKUP and XLOOKUP\n"
    "• Tasks: Step-by-step exercises\n"
    "• Hints & Answers: Check your work\n"
    "• Checklist: Tick off what you’ve completed"
)
ws["A3"].alignment = wrap

ws["A8"] = "Quick steps"
ws["A8"].font = header_font
ws["A9"] = (
    "1) Go to the Lookup sheet. Use the StudentID dropdown (cell B3).\n"
    "2) Enter VLOOKUP in cells B4 (Name) and B5 (Grade). Use exact match (FALSE) and lock the table with $.\n"
    "3) Try XLOOKUP in cells B7 (Name) and B8 (Grade). If your Excel doesn’t have XLOOKUP, skip this.\n"
    "4) Complete the Tasks sheet, then compare with Answers."
)
ws["A9"].alignment = wrap

ws["A14"] = "Tip: If copying formulas, make the table absolute like Data!$A$2:$D$11"
ws.column_dimensions["A"].width = 100

# ---------- Sheet: Data ----------
data_ws = wb.create_sheet("Data")

headers = ["StudentID", "Name", "Subject", "Grade"]
rows = [
    ["S101", "Amir", "Math", 85],
    ["S102", "Bella", "Math", 72],
    ["S103", "Chen", "Math", 91],
    ["S104", "Devi", "Math", 64],
    ["S105", "Ethan", "Math", 77],
    ["S106", "Farah", "Math", 88],
    ["S107", "Gino", "Math", 59],
    ["S108", "Hana", "Math", 95],
    ["S109", "Ivan", "Math", 73],
    ["S110", "Jade", "Math", 81],
]

# Write headers
for col, h in enumerate(headers, start=1):
    c = data_ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = fill_header
    c.alignment = center
    c.border = border_all

# Write data rows
for r, row in enumerate(rows, start=2):
    for c, val in enumerate(row, start=1):
        cell = data_ws.cell(row=r, column=c, value=val)
        if c == 4:
            # Grade formatting
            cell.number_format = "0"
        cell.border = border_all

# column widths
data_ws.column_dimensions["A"].width = 12
data_ws.column_dimensions["B"].width = 14
data_ws.column_dimensions["C"].width = 12
data_ws.column_dimensions["D"].width = 10

# Create a Table A1:D11
table_ref = "A1:D11"
tbl = Table(displayName="tblStudents", ref=table_ref)
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
tbl.tableStyleInfo = style
data_ws.add_table(tbl)

# Conditional formatting: highlight grades >= 85
rule = CellIsRule(operator="greaterThanOrEqual", formula=["85"])
# Use a simple 3-color scale for entire Grade column (D2:D11)
color_scale = ColorScaleRule(
    start_type="num",
    start_value=50,
    mid_type="num",
    mid_value=75,
    end_type="num",
    end_value=100,
)
data_ws.conditional_formatting.add("D2:D11", color_scale)

# Chart: Column chart of Grades by Name
chart = BarChart()
chart.title = "Grades by Student"
chart.y_axis.title = "Grade"
chart.x_axis.title = "Student"
cat = Reference(data_ws, min_col=2, min_row=2, max_row=11)  # Names
val = Reference(
    data_ws, min_col=4, min_row=1, max_row=11
)  # Include header for series name
chart.add_data(val, titles_from_data=True)
chart.set_categories(cat)
chart.height = 10
chart.width = 18
data_ws.add_chart(chart, "F2")

# ---------- Sheet: Lookup ----------
lk = wb.create_sheet("Lookup")
lk["A1"] = "Lookup a Student by ID"
lk["A1"].font = title_font
lk["A3"] = "StudentID:"
lk["A4"] = "Name (VLOOKUP):"
lk["A5"] = "Grade (VLOOKUP):"
lk["A7"] = "Name (XLOOKUP):"
lk["A8"] = "Grade (XLOOKUP):"
for r in [3, 4, 5, 7, 8]:
    lk.cell(row=r, column=1).font = header_font

# Data validation list for StudentID dropdown from Data sheet
dv = DataValidation(type="list", formula1="=Data!$A$2:$A$11", allow_blank=False)
lk.add_data_validation(dv)
dv.add(lk["B3"])

# Placeholder hints in right column
lk["D3"] = "Pick an ID from the dropdown."
lk["D4"] = "Enter VLOOKUP to return Name."
lk["D5"] = "Enter VLOOKUP to return Grade."
lk["D7"] = "Try XLOOKUP to return Name (if available)."
lk["D8"] = "Try XLOOKUP to return Grade."

# Pre-write example formulas as comments in cells below (not visible comments; just text helpers)
lk["A11"] = "VLOOKUP pattern:"
lk["B11"] = "=VLOOKUP(B3, Data!$A$2:$D$11, 2, FALSE)  → Name"
lk["B12"] = "=VLOOKUP(B3, Data!$A$2:$D$11, 4, FALSE)  → Grade"
lk["A14"] = "XLOOKUP pattern (Excel 365/2021+):"
lk["B14"] = "=XLOOKUP(B3, Data!$A$2:$A$11, Data!$B$2:$B$11)  → Name"
lk["B15"] = "=XLOOKUP(B3, Data!$A$2:$A$11, Data!$D$2:$D$11)  → Grade"

lk.column_dimensions["A"].width = 20
lk.column_dimensions["B"].width = 35
lk.column_dimensions["D"].width = 45

# ---------- Sheet: Tasks ----------
tasks = wb.create_sheet("Tasks")
tasks["A1"] = "Practice Tasks — Lookup Functions"
tasks["A1"].font = title_font

tasks_rows = [
    ["#", "Task", "Where", "Your Answer / Cell"],
    [1, "Use the dropdown to select StudentID S103.", "Lookup!B3", ""],
    [2, "Return the Name with VLOOKUP.", "Lookup!B4", ""],
    [3, "Return the Grade with VLOOKUP.", "Lookup!B5", ""],
    [
        4,
        "Copy your VLOOKUP to work for any selected ID (ensure $).",
        "Lookup!B4:B5",
        "",
    ],
    [5, "Try XLOOKUP for Name.", "Lookup!B7", ""],
    [6, "Try XLOOKUP for Grade.", "Lookup!B8", ""],
    [7, "On Data sheet, change Jade’s grade to 86. See chart update.", "Data!D11", ""],
    [8, "BONUS: Count how many students scored ≥ 80 using COUNTIF.", "Any cell", ""],
]
for r_idx, row in enumerate(tasks_rows, start=1):
    for c_idx, val in enumerate(row, start=1):
        cell = tasks.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = header_font
            cell.fill = fill_header
        cell.border = border_all
tasks.column_dimensions["A"].width = 5
tasks.column_dimensions["B"].width = 60
tasks.column_dimensions["C"].width = 18
tasks.column_dimensions["D"].width = 25

# ---------- Sheet: Hints ----------
hints = wb.create_sheet("Hints")
hints["A1"] = "Hints"
hints["A1"].font = title_font
hints["A3"] = (
    "VLOOKUP syntax: =VLOOKUP(lookup_value, table_array, col_index_num, [range_lookup])\n"
    "• lookup_value → Lookup!B3\n"
    "• table_array → Data!$A$2:$D$11  (lock with $)\n"
    "• col_index_num → 2 for Name, 4 for Grade\n"
    "• [range_lookup] → FALSE (exact match)\n\n"
    "XLOOKUP syntax: =XLOOKUP(lookup_value, lookup_array, return_array)\n"
    "• lookup_value → Lookup!B3\n"
    "• lookup_array → Data!$A$2:$A$11\n"
    "• return_array → Data!$B$2:$B$11 (Name) or $D$2:$D$11 (Grade)\n\n"
    'COUNTIF example (Task 8): =COUNTIF(Data!D2:D11, ">=80")'
)
hints.column_dimensions["A"].width = 110
hints["A3"].alignment = wrap

# ---------- Sheet: Answers ----------
ans = wb.create_sheet("Answers")
ans["A1"] = "Model Answers / Checks"
ans["A1"].font = title_font
ans["A3"] = "Enter these directly in the Lookup cells to check yourself:"
ans["A5"] = "Lookup!B4 (VLOOKUP Name)"
ans["B5"] = "=VLOOKUP(B3, Data!$A$2:$D$11, 2, FALSE)"
ans["A6"] = "Lookup!B5 (VLOOKUP Grade)"
ans["B6"] = "=VLOOKUP(B3, Data!$A$2:$D$11, 4, FALSE)"
ans["A8"] = "Lookup!B7 (XLOOKUP Name)"
ans["B8"] = (
    '=IFERROR(XLOOKUP(B3, Data!$A$2:$A$11, Data!$B$2:$B$11), "XLOOKUP not available")'
)
ans["A9"] = "Lookup!B8 (XLOOKUP Grade)"
ans["B9"] = (
    '=IFERROR(XLOOKUP(B3, Data!$A$2:$A$11, Data!$D$2:$D$11), "XLOOKUP not available")'
)
ans["A11"] = "Task 8 (COUNT of grades ≥ 80)"
ans["B11"] = '=COUNTIF(Data!D2:D11, ">=80")'
ans.column_dimensions["A"].width = 32
ans.column_dimensions["B"].width = 80

# ---------- Sheet: Checklist ----------
check = wb.create_sheet("Checklist")
check["A1"] = "Student Checklist"
check["A1"].font = title_font
items = [
    "Opened Lookup sheet and used the dropdown",
    "Built VLOOKUP for Name (exact match, correct column)",
    "Built VLOOKUP for Grade (exact match, correct column)",
    "Locked table with absolute references ($)",
    "Tried XLOOKUP (if available)",
    "Updated a grade and saw the chart change",
    "Completed COUNTIF bonus task",
]
check["A3"] = "Done?"
check["B3"] = "Task"
check["A3"].font = header_font
check["B3"].font = header_font
check["A3"].fill = fill_header
check["B3"].fill = fill_header
for i, text in enumerate(items, start=4):
    check.cell(row=i, column=1, value="No")  # change to Yes when done
    check.cell(row=i, column=2, value=text)
    check.cell(row=i, column=1).border = border_all
    check.cell(row=i, column=2).border = border_all
check.column_dimensions["A"].width = 8
check.column_dimensions["B"].width = 70

# Footer info
for ws_ in [data_ws, lk, tasks, hints, ans, check]:
    ws_["G100"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

# Save
wb.save("lookup_practice.xlsx")
print("Workbook created: lookup_practice.xlsx")
