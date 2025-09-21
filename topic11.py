# create_simple_data_analysis_workbook.py
# N Level Excel Starter: Simple Data Analysis (Percentages, Conditional Formatting, Charts)
# Creates sheets: Instructions, Data, Tasks, Hints, Answers, Checklist, Lookup

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
    NamedStyle,
    numbers,
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

# ---------- Helpers ----------
thin = Side(style="thin", color="CCCCCC")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def title(ws, text, cell="A1"):
    ws[cell] = text
    ws[cell].font = Font(size=16, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)


# ---------- Workbook & Sheets ----------
wb = Workbook()
ws_instr = wb.active
ws_instr.title = "Instructions"
ws_data = wb.create_sheet("Data")
ws_tasks = wb.create_sheet("Tasks")
ws_hints = wb.create_sheet("Hints")
ws_answers = wb.create_sheet("Answers")
ws_check = wb.create_sheet("Checklist")
ws_lookup = wb.create_sheet("Lookup")

# ---------- Instructions ----------
title(ws_instr, "Simple Data Analysis – Starter Workbook")
ws_instr["A3"] = "Objective:"
ws_instr["A3"].font = Font(bold=True)
ws_instr["B3"] = (
    "Calculate % change, share of total, and highlight trends with conditional formatting."
)

ws_instr["A5"] = "Skills covered:"
ws_instr["A5"].font = Font(bold=True)
ws_instr["B5"] = (
    "Formulas, absolute references ($), percentages, conditional formatting, chart creation, sorting/filtering."
)

ws_instr["A7"] = "Keyboard shortcuts (Windows / Mac):"
ws_instr["A7"].font = Font(bold=True)
ws_instr["B7"] = (
    "Copy: Ctrl+C / Cmd+C | Paste: Ctrl+V / Cmd+V | Fill down: Ctrl+D / Cmd+D"
)
ws_instr["B8"] = (
    "Format cells: Ctrl+1 / Cmd+1 | Create chart: Alt+N then pick chart / Cmd+Option+R (Excel menu)"
)

ws_instr["A10"] = "How to use:"
ws_instr["A10"].font = Font(bold=True)
ws_instr["B11"] = "1) Go to the Data sheet. Review sample products."
ws_instr["B12"] = "2) Enter or edit 2024 and 2025 sales."
ws_instr["B13"] = "3) Check formulas auto-filled in % Change and Share of Total."
ws_instr["B14"] = (
    "4) See conditional formatting highlight increases (green) and decreases (red)."
)
ws_instr["B15"] = "5) Explore the Tasks sheet, use Hints if stuck, then check Answers."
ws_instr["B16"] = "6) Use Checklist to self-assess."
ws_instr["B17"] = (
    "7) View the chart (Data sheet). Try changing the data and see it update."
)

set_col_widths(ws_instr, {"A": 22, "B": 100})
for r in range(3, 18):
    ws_instr[f"A{r}"].alignment = Alignment(vertical="top")
    ws_instr[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

# ---------- Lookup (for validation lists, references) ----------
title(ws_lookup, "Lookup & Reference")
ws_lookup["A3"] = "Categories (for Data validation):"
ws_lookup["A3"].font = Font(bold=True)
categories = ["Beverages", "Snacks", "Household", "Personal Care", "Electronics"]
for i, cat in enumerate(categories, start=4):
    ws_lookup[f"A{i}"] = cat
set_col_widths(ws_lookup, {"A": 28, "B": 60})

# ---------- Data (sample table + formulas + CF + chart) ----------
title(ws_data, "Sales Data (2024 vs 2025)")
headers = [
    "Product",
    "Category",
    "2024 Sales",
    "2025 Sales",
    "% Change",
    "Share of 2025 Total",
    "Status",
]
ws_data.append(headers)

sample_rows = [
    ["Cola 330ml", "Beverages", 1200, 1500],
    ["Orange Juice 1L", "Beverages", 980, 920],
    ["Potato Chips", "Snacks", 1500, 1800],
    ["Chocolate Bar", "Snacks", 1100, 1050],
    ["Detergent 2kg", "Household", 1350, 1600],
    ["Shampoo 500ml", "Personal Care", 900, 950],
    ["Earbuds", "Electronics", 2100, 2600],
    ["USB Charger", "Electronics", 1000, 900],
]

for r in sample_rows:
    ws_data.append(r + ["", "", ""])  # placeholders for % Change, Share, Status

# Headers style
for c in range(1, len(headers) + 1):
    cell = ws_data.cell(row=2, column=c)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="F2F2F2")
    cell.border = border_all
    cell.alignment = Alignment(horizontal="center")

# Freeze header row
ws_data.freeze_panes = "A3"

# Column widths
set_col_widths(ws_data, {"A": 22, "B": 18, "C": 14, "D": 14, "E": 12, "F": 20, "G": 12})

# Data range rows (after title row): headers at row 2, data rows 3..10
first_row = 3
last_row = first_row + len(sample_rows) - 1  # 10
total_row = last_row + 1  # 11

# Formulas
for r in range(first_row, last_row + 1):
    # % Change = IFERROR((New-Old)/Old,0)
    ws_data[f"E{r}"] = f"=IFERROR((D{r}-C{r})/C{r},0)"
    # Share of 2025 Total = IFERROR(D / SUM($D$first:$D$last),0)
    ws_data[f"F{r}"] = f"=IFERROR(D{r}/SUM($D${first_row}:$D${last_row}),0)"
    # Status text
    ws_data[f"G{r}"] = f'=IF(E{r}>0,"Increase",IF(E{r}<0,"Decrease","No change"))'

# Totals row
ws_data[f"A{total_row}"] = "Total"
ws_data[f"C{total_row}"] = f"=SUM(C{first_row}:C{last_row})"
ws_data[f"D{total_row}"] = f"=SUM(D{first_row}:D{last_row})"
ws_data[f"E{total_row}"] = ""  # leave blank
ws_data[f"F{total_row}"] = "1"  # total share = 100%
ws_data[f"G{total_row}"] = ""

# Number formats
for r in range(first_row, total_row + 1):
    ws_data[f"C{r}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    ws_data[f"D{r}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    ws_data[f"E{r}"].number_format = "0%"
    ws_data[f"F{r}"].number_format = "0%"

# Borders for data area
for r in range(2, total_row + 1):
    for c in range(1, len(headers) + 1):
        ws_data.cell(row=r, column=c).border = border_all

# Table
table_ref = f"A2:G{total_row}"
table = Table(displayName="tblSales", ref=table_ref)
style = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
)
table.tableStyleInfo = style
ws_data.add_table(table)

# Data Validation for Category (B column)
dv = DataValidation(
    type="list", formula1=f"=Lookup!$A$4:$A${3 + len(categories)}", allow_blank=False
)
dv.error = "Please select a category from the list."
dv.promptTitle = "Category"
dv.prompt = "Choose a category from Lookup sheet."
ws_data.add_data_validation(dv)
dv.add(f"B{first_row}:B{last_row}")

# Conditional Formatting on % Change (E)
ws_data.conditional_formatting.add(
    f"E{first_row}:E{last_row}",
    CellIsRule(
        operator="greaterThan",
        formula=["0"],
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    ),
)
ws_data.conditional_formatting.add(
    f"E{first_row}:E{last_row}",
    CellIsRule(
        operator="lessThan",
        formula=["0"],
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    ),
)

# Chart: Pie of 2025 Sales by Product
pie = PieChart()
labels = Reference(ws_data, min_col=1, min_row=first_row, max_row=last_row)  # Product
data = Reference(
    ws_data, min_col=4, min_row=2, max_row=last_row
)  # 2025 Sales incl header
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "2025 Sales Share"
pie.height = 12
pie.width = 18
ws_data.add_chart(pie, "I3")

# Chart: Column for 2024 vs 2025 by Product
bar = BarChart()
bar.title = "Sales by Product (2024 vs 2025)"
bar.height = 12
bar.width = 22
bar_data = Reference(
    ws_data, min_col=3, max_col=4, min_row=2, max_row=last_row
)  # 2024 & 2025
bar.add_data(bar_data, titles_from_data=True)
bar.set_categories(Reference(ws_data, min_col=1, min_row=first_row, max_row=last_row))
ws_data.add_chart(bar, "I20")

# ---------- Tasks ----------
title(ws_tasks, "Your Tasks")
tasks = [
    "Starter: Enter two new products at the bottom with 2024 & 2025 sales. Confirm % Change and Share fill automatically.",
    "Core: Apply a filter to show only 'Snacks'. Which product improved the most?",
    "Core: Sort by % Change (largest to smallest). Which 3 products increased the most?",
    "Stretch: Add a 'Target 2025 Sales' column (e.g., D * 1.10) and a 'Met Target?' column using IF.",
    "Stretch: Create a new pie chart of 2024 sales share.",
]
for i, t in enumerate(tasks, start=3):
    ws_tasks[f"A{i}"] = f"{i - 2}."
    ws_tasks[f"B{i}"] = t
set_col_widths(ws_tasks, {"A": 6, "B": 100})

# ---------- Hints ----------
title(ws_hints, "Hints")
hints = [
    "Percentage change = (New – Old) / Old. In this sheet: =(D - C) / C",
    "Share of total uses absolute refs: D / SUM($D$start:$D$end)",
    'IF example: =IF(E2>0,"Increase",IF(E2<0,"Decrease","No change"))',
    "To copy formulas, use the table fill handle or Ctrl+D (Cmd+D on Mac).",
    "Filter: Data tab → Filter (or click the ▼ on the table headers).",
    "Sort: Home → Sort & Filter → Sort Largest to Smallest on % Change.",
]
for i, h in enumerate(hints, start=3):
    ws_hints[f"A{i}"] = f"Hint {i - 2}"
    ws_hints[f"B{i}"] = h
set_col_widths(ws_hints, {"A": 12, "B": 100})

# ---------- Answers ----------
title(ws_answers, "Answers / Checks")
ws_answers["A3"] = "Key formulas used (check your sheet matches):"
ws_answers["A3"].font = Font(bold=True)
ws_answers["A5"] = "% Change (E row):"
ws_answers["B5"] = "=IFERROR((D2-C2)/C2,0)  → format as %"
ws_answers["A6"] = "Share of 2025 Total (F row):"
ws_answers["B6"] = f"=IFERROR(D2/SUM($D${first_row}:$D${last_row}),0)  → format as %"
ws_answers["A8"] = "Status (G row):"
ws_answers["B8"] = '=IF(E2>0,"Increase",IF(E2<0,"Decrease","No change"))'
ws_answers["A10"] = "Totals row:"
ws_answers["B10"] = (
    f"2024 Total =SUM(Data!C{first_row}:C{last_row}) | 2025 Total =SUM(Data!D{first_row}:D{last_row})"
)
ws_answers["A12"] = "Checks:"
ws_answers["B12"] = "Share column should sum to 100% (Total row shows 1.00)."
set_col_widths(ws_answers, {"A": 24, "B": 100})

# ---------- Checklist ----------
title(ws_check, "Checklist")
check_items = [
    "[ ] Entered/edited sales data for all rows",
    "[ ] % Change shows positives and negatives correctly",
    "[ ] Share of Total sums to 100%",
    "[ ] Conditional formatting highlights increases (green) and decreases (red)",
    "[ ] Applied sort/filter correctly",
    "[ ] Created and read the chart(s)",
    "[ ] Used absolute references ($) where needed",
]
for i, item in enumerate(check_items, start=3):
    ws_check[f"A{i}"] = item
set_col_widths(ws_check, {"A": 80})

# ---------- Finish ----------
# Make sheets user-friendly starting positions
for ws in [ws_data, ws_tasks, ws_hints, ws_answers, ws_check, ws_lookup]:
    ws.sheet_view.zoomScale = 120

# Save
wb.save("Simple_Data_Analysis_Starter.xlsx")
print("Workbook created: Simple_Data_Analysis_Starter.xlsx")
