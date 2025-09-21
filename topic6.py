# if_function_starter.py
# Creates an Excel workbook for N Level practice: IF function (basic logic)
# Sheets: Instructions, Data, Tasks, Hints, Answers, Checklist, Lookup
# Includes table formatting, data validation, and a simple chart.

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from datetime import datetime


# ---------- helpers ----------
def set_col_width(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def title(ws, text, cell="A1"):
    ws[cell] = text
    ws[cell].font = Font(size=16, bold=True)
    ws.merge_cells(f"{cell}:{get_column_letter(ws.max_column or 8)}{cell[1:]}")


def subhead(ws, row, text):
    ws[f"A{row}"] = text
    ws[f"A{row}"].font = Font(size=12, bold=True)


def add_table(ws, start_cell, end_cell, name, style="TableStyleMedium9"):
    table = Table(displayName=name, ref=f"{start_cell}:{end_cell}")
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


thin = Side(style="thin", color="CCCCCC")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="F2F2F2")

# ---------- workbook ----------
wb = Workbook()

# Remove default sheet
default = wb.active
wb.remove(default)

# ---------- Lookup (lists for validation etc.) ----------
wsL = wb.create_sheet("Lookup")
title(wsL, "Lookup Lists")
wsL["A3"] = "PassFailTexts"
wsL["A4"] = "Pass"
wsL["A5"] = "Fail"

wsL["C3"] = "DiscountTexts"
wsL["C4"] = "Discount"
wsL["C5"] = "No Discount"

wsL["E3"] = "GradeTexts"
wsL["E4"] = "A"
wsL["E5"] = "Pass"
wsL["E6"] = "Fail"

set_col_width(wsL, {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18})
wsL.freeze_panes = "A3"

# ---------- Instructions ----------
wsI = wb.create_sheet("Instructions")
set_col_width(wsI, {"A": 90})
title(wsI, "IF Function (Basic) — Practice Workbook")

wsI["A3"] = (
    "Objective: Use the IF function to make decisions in Excel (Pass/Fail, "
    "Discount flag, and simple grading with nested IF)."
)
wsI["A5"] = "How this workbook is organized:"
wsI["A6"] = "• Data: Sample records for marks, ages, and purchases."
wsI["A7"] = "• Tasks: Step-by-step activities (Starter → Core → Stretch)."
wsI["A8"] = "• Hints: Gentle nudges if you get stuck."
wsI["A9"] = "• Answers: Model answers and formulas to self-check."
wsI["A10"] = "• Checklist: Skills to tick off as you learn."
wsI["A12"] = "Keyboard tips (Windows / Mac):"
wsI["A13"] = "• Edit cell: F2 / Control+U"
wsI["A14"] = "• Fill down: Ctrl+D / Command+D"
wsI["A15"] = "• Fill right: Ctrl+R / Command+R"
wsI["A16"] = "• Create table: Ctrl+T / Command+T"

wsI["A18"] = 'Reminder: Text results like Pass/Fail must be inside quotes, e.g. "Pass".'

# ---------- Data ----------
wsD = wb.create_sheet("Data")
title(wsD, "Practice Data")
headers = [
    "Name",
    "Age",
    "Exam Mark",
    "Purchase ($)",
    "Pass/Fail",
    "Discount?",
    "Grade",
]
wsD.append(headers)

data_rows = [
    ["Aiden", 17, 82, 120, "", "", ""],
    ["Bella", 19, 47, 95, "", "", ""],
    ["Chloe", 21, 50, 205, "", "", ""],
    ["Darius", 16, 73, 40, "", "", ""],
    ["Eli", 18, 33, 100, "", "", ""],
    ["Farah", 20, 89, 155, "", "", ""],
    ["Gwen", 22, 58, 60, "", "", ""],
    ["Hugo", 17, 79, 99, "", "", ""],
    ["Iris", 18, 51, 300, "", "", ""],
    ["Jules", 23, 45, 110, "", "", ""],
]
for r in data_rows:
    wsD.append(r)

# style header
for col in range(1, len(headers) + 1):
    cell = wsD.cell(row=2, column=col)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.border = border_all
    wsD.cell(row=2, column=col).alignment = Alignment(
        horizontal="center", vertical="center"
    )

# borders for data
for r in range(3, 3 + len(data_rows)):
    for c in range(1, len(headers) + 1):
        wsD.cell(row=r, column=c).border = border_all

set_col_width(wsD, {"A": 14, "B": 8, "C": 10, "D": 12, "E": 12, "F": 12, "G": 10})
wsD.freeze_panes = "A3"

# turn into a table
last_row = 2 + len(data_rows) + 1
add_table(wsD, "A2", f"G{2 + len(data_rows)}", "tblData")

# Data Validations (optional dropdowns for checking)
dv_passfail = DataValidation(
    type="list", formula1="=Lookup!$A$4:$A$5", allow_blank=True
)
dv_discount = DataValidation(
    type="list", formula1="=Lookup!$C$4:$C$5", allow_blank=True
)
dv_grade = DataValidation(type="list", formula1="=Lookup!$E$4:$E$6", allow_blank=True)
wsD.add_data_validation(dv_passfail)
wsD.add_data_validation(dv_discount)
wsD.add_data_validation(dv_grade)
dv_passfail.add(f"E3:E{2 + len(data_rows)}")
dv_discount.add(f"F3:F{2 + len(data_rows)}")
dv_grade.add(f"G3:G{2 + len(data_rows)}")

# Simple bar chart: Exam Mark by Name
chart = BarChart()
chart.title = "Exam Marks"
chart.y_axis.title = "Mark"
chart.x_axis.title = "Name"

data_ref = Reference(wsD, min_col=3, min_row=2, max_row=2 + len(data_rows))
cats_ref = Reference(wsD, min_col=1, min_row=3, max_row=2 + len(data_rows))
chart.add_data(data_ref, from_rows=False, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 11
chart.width = 20
wsD.add_chart(chart, "I3")

# ---------- Tasks ----------
wsT = wb.create_sheet("Tasks")
title(wsT, "Tasks — Starter → Core → Stretch")
set_col_width(wsT, {"A": 90})
wsT["A3"] = "Starter (IF basics):"
wsT["A4"] = (
    'In Data!E3, write an IF formula to show "Pass" if Exam Mark (column C) ≥ 50, '
    'otherwise "Fail". Fill down to E12.'
)
wsT["A6"] = "Core (another IF):"
wsT["A7"] = (
    'In Data!F3, write an IF formula to show "Discount" if Purchase (column D) ≥ 100, '
    'otherwise "No Discount". Fill down to F12.'
)
wsT["A9"] = "Stretch (nested IF grading):"
wsT["A10"] = (
    'In Data!G3, write a nested IF: if Exam Mark ≥ 80 return "A"; else if Exam Mark ≥ 50 return "Pass"; '
    'otherwise return "Fail". Fill down to G12.'
)
wsT["A12"] = "Bonus (absolute reference practice):"
wsT["A13"] = (
    "Type the pass mark (50) in H3 and the discount threshold (100) in H4 on the Data sheet. "
    "Rewrite your formulas using absolute references to those cells (e.g., $H$3, $H$4)."
)

# ---------- Hints ----------
wsH = wb.create_sheet("Hints")
title(wsH, "Hints")
set_col_width(wsH, {"A": 90})
wsH["A3"] = "IF structure: =IF(condition, value_if_true, value_if_false)"
wsH["A5"] = 'Starter hint: =IF(C3>=50,"Pass","Fail")'
wsH["A7"] = 'Core hint: =IF(D3>=100,"Discount","No Discount")'
wsH["A9"] = 'Stretch hint (nested): =IF(C3>=80,"A",IF(C3>=50,"Pass","Fail"))'
wsH["A11"] = (
    "Absolute reference: Put 50 in Data!H3 and 100 in Data!H4, then use $H$3 and $H$4."
)
wsH["A12"] = 'Example: =IF(C3>=$H$3,"Pass","Fail")'
wsH["A14"] = "Text needs quotes. Numbers do not."
wsH["A15"] = (
    "Regional settings: If your Excel uses semicolons, replace commas with semicolons."
)

# ---------- Answers ----------
wsA = wb.create_sheet("Answers")
title(wsA, "Answers (Formulas)")
headers_ans = ["Task", "Cell", "Formula"]
wsA.append(headers_ans)
answers = [
    ["Starter — Pass/Fail", "Data!E3", '=IF(C3>=50,"Pass","Fail")'],
    ["Core — Discount flag", "Data!F3", '=IF(D3>=100,"Discount","No Discount")'],
    [
        "Stretch — Grade (nested IF)",
        "Data!G3",
        '=IF(C3>=80,"A",IF(C3>=50,"Pass","Fail"))',
    ],
    ["Bonus — Pass/Fail w/ $", "Data!E3", '=IF(C3>=$H$3,"Pass","Fail")'],
    ["Bonus — Discount w/ $", "Data!F3", '=IF(D3>=$H$4,"Discount","No Discount")'],
]
for row in answers:
    wsA.append(row)

# style header
for col in range(1, len(headers_ans) + 1):
    cell = wsA.cell(row=2, column=col)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.border = border_all
for r in range(3, 3 + len(answers)):
    for c in range(1, len(headers_ans) + 1):
        wsA.cell(row=r, column=c).border = border_all

set_col_width(wsA, {"A": 28, "B": 14, "C": 60})
add_table(wsA, "A2", f"C{2 + len(answers)}", "tblAnswers")

# ---------- Checklist ----------
wsC = wb.create_sheet("Checklist")
title(wsC, "Checklist — Tick as you complete")
set_col_width(wsC, {"A": 60, "B": 14})
wsC.append(["Skill", "Done? (Y/N)"])
check_items = [
    "Typed a basic IF formula",
    "Used comparison operators (>=, <)",
    "Filled a formula down a column",
    "Nested an IF inside another IF",
    "Used absolute references ($H$3, $H$4)",
    "Created/used a Table (Ctrl+T / Command+T)",
    "Understood quotes for text vs numbers",
]
for item in check_items:
    wsC.append([item, ""])
add_table(wsC, "A2", f"B{2 + len(check_items)}", "tblChecklist")

# ---------- Finishing touches ----------
# Put threshold placeholders in Data for bonus task
wsD["H2"] = "Thresholds"
wsD["H3"] = 50
wsD["H4"] = 100
wsD["H2"].font = Font(bold=True)
wsD["H2"].fill = header_fill
wsD["H3"].number_format = "0"
wsD["H4"].number_format = "0"
wsD["H2"].border = wsD["H3"].border = wsD["H4"].border = border_all

# Footer notes
wsI["A20"] = f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

# Save
wb.save("IF_Function_Starter.xlsx")
print("Workbook created: IF_Function_Starter.xlsx")
