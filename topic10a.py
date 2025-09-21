# create_sort_filter_workbook.py
# Creates: Sorting_Filtering_Practice.xlsx
# Requires: pip install openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

# -----------------------------
# Helper styles
# -----------------------------
header_fill = PatternFill("solid", fgColor="DDEAF6")
thin = Side(style="thin", color="999999")
thin_border = Border(top=thin, bottom=thin, left=thin, right=thin)

title_style = NamedStyle(name="title_style")
title_style.font = Font(b=True, size=14)
title_style.alignment = Alignment(vertical="center")
if "title_style" not in wb.named_styles:
    wb.add_named_style(title_style)

hdr_style = NamedStyle(name="hdr_style")
hdr_style.font = Font(b=True)
hdr_style.fill = header_fill
hdr_style.border = thin_border
hdr_style.alignment = Alignment(horizontal="center", vertical="center")
if "hdr_style" not in wb.named_styles:
    wb.add_named_style(hdr_style)

normal_num = NamedStyle(name="normal_num")
normal_num.number_format = "#,##0"
if "normal_num" not in wb.named_styles:
    wb.add_named_style(normal_num)

currency_style = NamedStyle(name="currency_style")
currency_style.number_format = '"$"#,##0.00'
if "currency_style" not in wb.named_styles:
    wb.add_named_style(currency_style)

date_style = NamedStyle(name="date_style")
date_style.number_format = "yyyy-mm-dd"
if "date_style" not in wb.named_styles:
    wb.add_named_style(date_style)

# -----------------------------
# Sheets
# -----------------------------
ws_instr = wb.active
ws_instr.title = "Instructions"
ws_data = wb.create_sheet("Data")
ws_tasks = wb.create_sheet("Tasks")
ws_hints = wb.create_sheet("Hints")
ws_answers = wb.create_sheet("Answers")
ws_check = wb.create_sheet("Checklist")
ws_lookup = wb.create_sheet("Lookup")

# -----------------------------
# Instructions sheet
# -----------------------------
ws = ws_instr
ws["A1"] = "Excel Practice: Sorting & Filtering"
ws["A1"].style = "title_style"
ws["A3"] = "How to use this workbook"
ws["A3"].font = Font(b=True)

instr_lines = [
    "1) Go to the Data sheet. The sales table already has Filter drop-downs.",
    "2) Complete each task on the Tasks sheet by performing the action on the Data table.",
    "3) Check Hints if you’re stuck. Compare with the Answers sheet to self-check.",
    "4) Use Ctrl + Z to undo. Don’t type into the Data table except Units/Price (if exploring).",
    "",
    "Shortcuts:",
    "• Toggle Filters: Ctrl + Shift + L",
    "• Go to Data tab: Alt, A (Windows) / Use Ribbon on Mac",
    "• Sort A→Z / Z→A from column filter menus or Data tab",
]
for i, t in enumerate(instr_lines, start=4):
    ws[f"A{i}"] = t

# quick nav links (Excel turns these into clickable links in many viewers)
ws["A12"] = "Open Data →"
ws["A12"].hyperlink = "#'Data'!A1"
ws["A12"].style = "Hyperlink"

ws["A13"] = "Open Tasks →"
ws["A13"].hyperlink = "#'Tasks'!A1"
ws["A13"].style = "Hyperlink"

ws.column_dimensions["A"].width = 95

# -----------------------------
# Lookup sheet (for validation)
# -----------------------------
ws = ws_lookup
ws["A1"] = "Region"
regions = ["East", "West", "North", "South"]
for r, val in enumerate(regions, start=2):
    ws[f"A{r}"] = val
ws.column_dimensions["A"].width = 18

# -----------------------------
# Data sheet with a Table
# -----------------------------
ws = ws_data
headers = [
    "Order ID",
    "Date",
    "Name",
    "Region",
    "Product",
    "Units",
    "Unit Price",
    "Sales",
]
ws.append(headers)

data_rows = [
    [1001, "2025-02-02", "Alex", "East", "Notebook", 12, 4.50],
    [1002, "2025-02-05", "Bella", "West", "Binder", 15, 6.20],
    [1003, "2025-02-07", "Chris", "East", "Pen", 50, 1.20],
    [1004, "2025-02-08", "Diana", "North", "Pencil", 60, 0.80],
    [1005, "2025-02-09", "Evan", "South", "Notebook", 20, 4.50],
    [1006, "2025-02-12", "Fiona", "West", "Binder", 35, 6.20],
    [1007, "2025-02-14", "Gina", "North", "Pen", 70, 1.20],
    [1008, "2025-02-17", "Henry", "East", "Pencil", 80, 0.80],
    [1009, "2025-02-20", "Iris", "South", "Notebook", 18, 4.50],
    [1010, "2025-02-22", "Jack", "West", "Pen", 90, 1.20],
    [1011, "2025-02-25", "Kara", "North", "Binder", 12, 6.20],
    [1012, "2025-02-27", "Liam", "East", "Notebook", 25, 4.50],
    [1013, "2025-03-01", "Maya", "South", "Pencil", 55, 0.80],
    [1014, "2025-03-03", "Noah", "West", "Notebook", 22, 4.50],
    [1015, "2025-03-05", "Olive", "North", "Pen", 65, 1.20],
    [1016, "2025-03-08", "Paul", "East", "Binder", 28, 6.20],
]

start_row = 2
for i, row in enumerate(data_rows, start=start_row):
    # Append base fields (A:G)
    ws.append(row + [None])  # placeholder for Sales in H
    # Apply styles
    ws[f"A{i}"].style = normal_num
    ws[f"B{i}"].style = "date_style"
    ws[f"F{i}"].style = normal_num
    ws[f"G{i}"].style = "currency_style"
    # Sales formula = Units * Unit Price
    ws[f"H{i}"] = f"=F{i}*G{i}"
    ws[f"H{i}"].style = "currency_style"

# Header styling
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.value = headers[col - 1]
    c.style = "hdr_style"
    ws.column_dimensions[get_column_letter(col)].width = [
        12,
        12,
        14,
        12,
        14,
        10,
        12,
        12,
    ][col - 1]

# Data validation for Region (D column) using Lookup sheet A2:A5
dv = DataValidation(
    type="list", formula1="=Lookup!$A$2:$A$5", allow_blank=False, showDropDown=True
)
ws.add_data_validation(dv)
dv.add(f"D{start_row}:D{start_row + len(data_rows) - 1}")

# Freeze panes and filter via Table
ws.freeze_panes = "A2"

# Create an Excel Table with filters
last_row = start_row + len(data_rows) - 1
table_ref = f"A1:H{last_row}"
table = Table(displayName="SalesData", ref=table_ref)
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
table.tableStyleInfo = style
ws.add_table(table)

# Region totals (for chart) in columns K:L
ws["K1"] = "Region"
ws["L1"] = "Total Sales"
ws["K1"].style = "hdr_style"
ws["L1"].style = "hdr_style"
for idx, reg in enumerate(regions, start=2):
    ws[f"K{idx}"] = reg
    ws[f"L{idx}"] = f"=SUMIF($D$2:$D${last_row}, K{idx}, $H$2:$H${last_row})"
    ws[f"L{idx}"].style = "currency_style"
ws.column_dimensions["K"].width = 12
ws.column_dimensions["L"].width = 14

# Chart (Column chart of sales by region)
chart = BarChart()
chart.title = "Total Sales by Region"
chart.y_axis.title = "Sales ($)"
chart.x_axis.title = "Region"
data = Reference(ws, min_col=12, min_row=1, max_row=1 + len(regions))  # L1:L5
cats = Reference(ws, min_col=11, min_row=2, max_row=1 + len(regions))  # K2:K5
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 9
chart.width = 17
ws.add_chart(chart, "N2")

# -----------------------------
# Tasks sheet
# -----------------------------
ws = ws_tasks
ws["A1"] = "Tasks: Sorting & Filtering (work on the Data sheet)"
ws["A1"].style = "title_style"
tasks = [
    "Task 1 — Sort Sales from highest to lowest (Z→A on Sales).",
    "Task 2 — Sort Names A→Z.",
    "Task 3 — Filter to show only Region = West.",
    "Task 4 — Combine: Filter Region = East, then sort Sales Z→A.",
    "Bonus — Clear filters and sort by Date oldest→newest.",
]
for i, t in enumerate(tasks, start=3):
    ws[f"A{i}"] = t

ws["A9"] = (
    "Tip: Perform the actions directly on the table in the Data sheet. Use Answers sheet to self-check."
)
ws["A11"] = "Open Data →"
ws["A11"].hyperlink = "#'Data'!A1"
ws["A11"].style = "Hyperlink"
ws.column_dimensions["A"].width = 100

# -----------------------------
# Hints sheet
# -----------------------------
ws = ws_hints
ws["A1"] = "Hints"
ws["A1"].style = "title_style"
hint_lines = [
    "Sorting:",
    "• Click any cell in the column you want to sort (e.g., Sales).",
    "• Home → Sort & Filter → Sort Largest to Smallest (or Data tab → Sort Z→A).",
    "",
    "Filtering:",
    "• Data → Filter (or Ctrl + Shift + L).",
    "• Click the drop-down in the Region header → (Select All) off → tick the region you want.",
    "• To remove: Open the same menu → Clear Filter from 'Region'.",
    "",
    "Combining:",
    "• You can filter first, then sort within the filtered rows.",
]
for i, t in enumerate(hint_lines, start=3):
    ws[f"A{i}"] = t
ws.column_dimensions["A"].width = 95

# -----------------------------
# Answers sheet (expected outcome tables)
# -----------------------------
ws = ws_answers
ws["A1"] = "Answers (Expected Results)"
ws["A1"].style = "title_style"


# Helper: make a small table printer
def write_table(ws, start_cell, title, header_row, rows):
    ws[start_cell] = title
    ws[start_cell].font = Font(b=True)
    col_start = start_cell[0]
    row_start = int(start_cell[1:])
    # header
    for j, h in enumerate(header_row, start=0):
        cell = ws.cell(row=row_start + 1, column=j + 1)
        cell.value = h
        cell.style = "hdr_style"
    # rows
    for i, r in enumerate(rows, start=0):
        for j, val in enumerate(r, start=0):
            cell = ws.cell(row=row_start + 2 + i, column=j + 1)
            cell.value = val
            if header_row[j] in ("Units", "Order ID"):
                cell.style = normal_num
            if header_row[j] == "Unit Price":
                cell.style = currency_style


# Pre-computed views based on the fixed data above


# Task 1: Sort by Sales desc (we'll compute Sales values to list correctly)
def sales(units, price):
    return round(units * price, 2)


full = [
    [1001, "2025-02-02", "Alex", "East", "Notebook", 12, 4.50, sales(12, 4.50)],
    [1002, "2025-02-05", "Bella", "West", "Binder", 15, 6.20, sales(15, 6.20)],
    [1003, "2025-02-07", "Chris", "East", "Pen", 50, 1.20, sales(50, 1.20)],
    [1004, "2025-02-08", "Diana", "North", "Pencil", 60, 0.80, sales(60, 0.80)],
    [1005, "2025-02-09", "Evan", "South", "Notebook", 20, 4.50, sales(20, 4.50)],
    [1006, "2025-02-12", "Fiona", "West", "Binder", 35, 6.20, sales(35, 6.20)],
    [1007, "2025-02-14", "Gina", "North", "Pen", 70, 1.20, sales(70, 1.20)],
    [1008, "2025-02-17", "Henry", "East", "Pencil", 80, 0.80, sales(80, 0.80)],
    [1009, "2025-02-20", "Iris", "South", "Notebook", 18, 4.50, sales(18, 4.50)],
    [1010, "2025-02-22", "Jack", "West", "Pen", 90, 1.20, sales(90, 1.20)],
    [1011, "2025-02-25", "Kara", "North", "Binder", 12, 6.20, sales(12, 6.20)],
    [1012, "2025-02-27", "Liam", "East", "Notebook", 25, 4.50, sales(25, 4.50)],
    [1013, "2025-03-01", "Maya", "South", "Pencil", 55, 0.80, sales(55, 0.80)],
    [1014, "2025-03-03", "Noah", "West", "Notebook", 22, 4.50, sales(22, 4.50)],
    [1015, "2025-03-05", "Olive", "North", "Pen", 65, 1.20, sales(65, 1.20)],
    [1016, "2025-03-08", "Paul", "East", "Binder", 28, 6.20, sales(28, 6.20)],
]

# Sort by Sales desc
task1_sorted = sorted(full, key=lambda r: r[7], reverse=True)

# Task 2: Names A→Z (stable sort by Name asc)
task2_sorted = sorted(full, key=lambda r: r[2])

# Task 3: Filter Region = West
task3_west = [r for r in full if r[3] == "West"]

# Task 4: Filter East then Sales desc
task4_east_sales = sorted(
    [r for r in full if r[3] == "East"], key=lambda r: r[7], reverse=True
)

header_full = [
    "Order ID",
    "Date",
    "Name",
    "Region",
    "Product",
    "Units",
    "Unit Price",
    "Sales",
]

write_table(
    ws,
    "A3",
    "Task 1 — Sales Z→A (Expected Order Top 10 Shown)",
    header_full,
    task1_sorted[:10],
)
write_table(
    ws, "A18", "Task 2 — Names A→Z (First 10 Shown)", header_full, task2_sorted[:10]
)
write_table(ws, "A33", "Task 3 — Region = West", header_full, task3_west)
write_table(
    ws, "A49", "Task 4 — Region = East then Sales Z→A", header_full, task4_east_sales
)

for col in range(1, 9):
    ws.column_dimensions[get_column_letter(col)].width = [
        10,
        12,
        12,
        10,
        12,
        8,
        12,
        12,
    ][col - 1]

# -----------------------------
# Checklist sheet
# -----------------------------
ws = ws_check
ws["A1"] = "Checklist — tick when done"
ws["A1"].style = "title_style"
check_items = [
    "[ ] I can turn Filters on/off (Ctrl + Shift + L).",
    "[ ] I can sort a numeric column Z→A and A→Z.",
    "[ ] I can sort a text column A→Z and Z→A.",
    "[ ] I can filter to a single Region.",
    "[ ] I can combine filter + sort.",
    "[ ] I can clear filters to show all rows.",
]
for i, t in enumerate(check_items, start=3):
    ws[f"A{i}"] = t
ws.column_dimensions["A"].width = 80

# -----------------------------
# Finishing touches
# -----------------------------
# Add simple borders to Instructions lists
for r in range(4, 4 + len(instr_lines)):
    ws_instr[f"A{r}"].border = thin_border

# Save
wb.save("Sorting_Filtering_Practice.xlsx")
print("Created Sorting_Filtering_Practice.xlsx")
