"""
Excel Starter Workbook Generator — Core Functions

Creates an .xlsx file with these sheets:
- Instructions (lesson recap + tasks overview)
- Data (sample Sales & Scores tables)
- Tasks (cells for students to enter formulas)
- Hints (gentle guidance)
- Answers (completed formulas for checking)
- Checklist (skills tick-off)
- Lookup (quick function reference)

Also adds:
- Excel Tables for Data ranges
- A simple column chart: Sales Amount by Item

How to run (Windows/Mac):
1) Ensure Python 3.8+ is installed.
2) Install openpyxl:  pip install openpyxl
3) Save this script as generate_core_functions_workbook.py
4) Run:  python generate_core_functions_workbook.py
5) The file "Core_Functions_Practice.xlsx" will be created in the same folder.

Note: This chat can’t auto-download files. Run locally to generate the workbook.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# -----------------------------
# Helper functions
# -----------------------------


def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

currency_fmt = "#,##0.00"

wb = Workbook()

# Remove default sheet name and start fresh
ws0 = wb.active
ws0.title = "Instructions"

# -----------------------------
# Sheet: Instructions
# -----------------------------
ws = wb["Instructions"]
ws.sheet_properties.tabColor = "38B6FF"  # blue

ws["A1"] = "Core Functions Practice: SUM, AVERAGE, MIN, MAX, COUNT, COUNTA"
ws["A1"].font = Font(size=14, bold=True)
ws.merge_cells("A1:F1")

text = (
    "Objective: Use basic summary functions to analyse data quickly.\n\n"
    "Why it matters: Formulas update automatically when data changes, saving time and reducing errors.\n\n"
    "What to do:\n"
    "1) Go to the Data sheet and review the Sales and Scores tables.\n"
    "2) On the Tasks sheet, enter formulas in the yellow cells.\n"
    "3) Use the Hints sheet if stuck.\n"
    "4) Check yourself with the Answers sheet.\n"
    "5) Tick items on the Checklist when done.\n\n"
    "Key functions:\n"
    "- SUM(range): Adds numbers.\n"
    "- AVERAGE(range): Mean value.\n"
    "- MIN(range): Smallest number.\n"
    "- MAX(range): Largest number.\n"
    "- COUNT(range): Counts numbers only.\n"
    "- COUNTA(range): Counts non-blank cells (numbers + text).\n"
)
ws["A3"] = text
ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

set_col_widths(ws, {"A": 90})
ws.row_dimensions[1].height = 24
ws.freeze_panes = "A4"

# -----------------------------
# Sheet: Data
# -----------------------------
ws = wb.create_sheet("Data")
ws.sheet_properties.tabColor = "92D050"  # green

# Sales table headers
ws["A2"].value = "Date"
ws["B2"].value = "Item"
ws["C2"].value = "Qty"
ws["D2"].value = "Unit Price"
ws["E2"].value = "Amount"

sales_rows = [
    ["2025-01-05", "Notebooks", 5, 2.5],
    ["2025-01-06", "Pencils", 20, 0.6],
    ["2025-01-06", "Erasers", 8, 0.8],
    ["2025-01-08", "Folders", 10, 1.2],
    ["2025-01-09", "Markers", 6, 1.5],
    ["2025-01-10", "Highlighters", 4, 1.8],
    ["2025-01-11", "Glue Sticks", 7, 1.1],
    ["2025-01-12", "Scissors", 3, 3.2],
    ["2025-01-13", "Staplers", 2, 5.0],
    ["2025-01-14", "Tape", 9, 1.0],
    ["2025-01-15", "Clip Sets", 6, 1.3],
    ["2025-01-16", "Rulers", 12, 0.9],
]

start_row = 3
for i, (date, item, qty, price) in enumerate(sales_rows, start=start_row):
    ws.cell(row=i, column=1, value=date)
    ws.cell(row=i, column=2, value=item)
    ws.cell(row=i, column=3, value=qty)
    p = ws.cell(row=i, column=4, value=price)
    p.number_format = currency_fmt
    # Amount formula = Qty * Unit Price
    amt = ws.cell(row=i, column=5)
    amt.value = f"=C{i}*D{i}"
    amt.number_format = currency_fmt

# Style header row
for col in range(1, 6):
    cell = ws.cell(row=2, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cell.border = thin_border

# Create Table for Sales
last_row = start_row + len(sales_rows) - 1
sales_table = Table(displayName="SalesTbl", ref=f"A2:E{last_row}")
sales_style = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
)
sales_table.tableStyleInfo = sales_style
ws.add_table(sales_table)

# Scores table headers (placed to the right)
ws["G2"].value = "Name"
ws["H2"].value = "Score"

scores_rows = [
    ["Ali", 65],
    ["Ben", 72],
    ["Clara", 80],
    ["Devi", 55],
    ["Emma", 90],
    ["Farah", 77],
    ["Gopal", 68],
    ["Hui", 83],
    ["Ivan", 59],
    ["Jin", 74],
]

for i, (name, score) in enumerate(scores_rows, start=start_row):
    ws.cell(row=i, column=7, value=name)
    ws.cell(row=i, column=8, value=score)

for col in (7, 8):
    cell = ws.cell(row=2, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    cell.border = thin_border

scores_last_row = start_row + len(scores_rows) - 1
scores_table = Table(displayName="ScoresTbl", ref=f"G2:H{scores_last_row}")
scores_style = TableStyleInfo(
    name="TableStyleMedium7", showRowStripes=True, showColumnStripes=False
)
scores_table.tableStyleInfo = scores_style
ws.add_table(scores_table)

# Column widths and freeze
set_col_widths(ws, {"A": 12, "B": 16, "C": 8, "D": 12, "E": 12, "G": 14, "H": 10})
ws.freeze_panes = "A3"

# Add a simple column chart: Sales Amount by Item
chart = BarChart()
chart.title = "Sales Amount by Item"
chart.y_axis.title = "Amount"
chart.x_axis.title = "Item"

amounts = Reference(ws, min_col=5, min_row=2, max_row=last_row)
items = Reference(ws, min_col=2, min_row=3, max_row=last_row)
chart.add_data(amounts, titles_from_data=True)
chart.set_categories(items)
ws.add_chart(chart, "A16")

# -----------------------------
# Sheet: Tasks
# -----------------------------
ws = wb.create_sheet("Tasks")
ws.sheet_properties.tabColor = "FFD966"  # yellow

set_col_widths(ws, {"A": 4, "B": 55, "C": 22})

ws["B2"].value = "Enter your formulas in the yellow cells (C column)."
ws["B2"].font = Font(bold=True)

tasks = [
    ("Total Sales (SUM of Data!E3:E14)", "=SUM(Data!E3:E14)"),
    ("Average Sale per order (AVERAGE of Data!E3:E14)", "=AVERAGE(Data!E3:E14)"),
    ("Smallest sale amount (MIN of Data!E3:E14)", "=MIN(Data!E3:E14)"),
    ("Largest sale amount (MAX of Data!E3:E14)", "=MAX(Data!E3:E14)"),
    ("Count of numeric scores (COUNT of Data!H3:H12)", "=COUNT(Data!H3:H12)"),
    ("Count of names (COUNTA of Data!G3:G12)", "=COUNTA(Data!G3:G12)"),
    ("BONUS: Total Quantity sold (SUM of Data!C3:C14)", "=SUM(Data!C3:C14)"),
]

start = 4
for i, (label, answer_formula) in enumerate(tasks, start=start):
    ws.cell(row=i, column=2, value=f"{i - start + 1}) {label}")
    target = ws.cell(row=i, column=3)
    target.value = None  # student to enter
    target.number_format = currency_fmt if i in (4, 5, 6, 7) else "General"
    target.fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    target.border = thin_border

ws["B12"] = (
    "Tip: Use = to start every formula. Select the correct range, including the last row."
)
ws["B12"].alignment = Alignment(wrap_text=True)

# -----------------------------
# Sheet: Hints
# -----------------------------
ws = wb.create_sheet("Hints")
ws.sheet_properties.tabColor = "B4A7D6"  # purple
set_col_widths(ws, {"A": 95})

hints = [
    "SUM adds numbers: =SUM(Data!E3:E14)",
    "AVERAGE finds the mean: =AVERAGE(Data!E3:E14)",
    "MIN gives the smallest value: =MIN(Data!E3:E14)",
    "MAX gives the largest value: =MAX(Data!E3:E14)",
    "COUNT counts numbers only: =COUNT(Data!H3:H12)",
    "COUNTA counts non-blank cells: =COUNTA(Data!G3:G12)",
    "Bonus idea: Total Qty =SUM(Data!C3:C14)",
]

ws["A1"].value = "Hints"
ws["A1"].font = Font(bold=True)
for i, line in enumerate(hints, start=3):
    ws.cell(row=i, column=1, value=f"• {line}")

# -----------------------------
# Sheet: Answers
# -----------------------------
ws = wb.create_sheet("Answers")
ws.sheet_properties.tabColor = "F4CCCC"  # red
set_col_widths(ws, {"A": 4, "B": 55, "C": 22})

ws["B2"].value = "Model answers (formulas are entered for you):"
ws["B2"].font = Font(bold=True)

for i, (label, formula) in enumerate(tasks, start=4):
    ws.cell(row=i, column=2, value=f"{i - 3}) {label}")
    ans = ws.cell(row=i, column=3, value=formula)
    ans.number_format = currency_fmt if i in (4, 5, 6, 7) else "General"
    ans.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ans.border = thin_border

# -----------------------------
# Sheet: Checklist
# -----------------------------
ws = wb.create_sheet("Checklist")
ws.sheet_properties.tabColor = "A2C4C9"  # teal
set_col_widths(ws, {"A": 70, "B": 18})

items = [
    "I can use =SUM(range) to add numbers.",
    "I can use =AVERAGE(range) to find the mean.",
    "I can identify the smallest and largest values using MIN and MAX.",
    "I know the difference between COUNT (numbers) and COUNTA (non-blanks).",
    "I can select the correct range, including the last row.",
]

ws["A1"].value = "Skill"
ws["B1"].value = "Done [Y/N]"
ws["A1"].font = Font(bold=True)
ws["B1"].font = Font(bold=True)

for i, item in enumerate(items, start=2):
    ws.cell(row=i, column=1, value=item)
    ws.cell(row=i, column=2, value="[ ]")

# -----------------------------
# Sheet: Lookup
# -----------------------------
ws = wb.create_sheet("Lookup")
ws.sheet_properties.tabColor = "CCCCCC"
set_col_widths(ws, {"A": 22, "B": 80, "C": 44})

lookup_rows = [
    ("Function", "Meaning / Syntax", "Example"),
    ("SUM", "Adds numbers — SUM(range)", "=SUM(Data!E3:E14)"),
    ("AVERAGE", "Mean value — AVERAGE(range)", "=AVERAGE(Data!E3:E14)"),
    ("MIN", "Smallest number — MIN(range)", "=MIN(Data!E3:E14)"),
    ("MAX", "Largest number — MAX(range)", "=MAX(Data!E3:E14)"),
    ("COUNT", "Counts numbers only — COUNT(range)", "=COUNT(Data!H3:H12)"),
    ("COUNTA", "Counts non-blanks — COUNTA(range)", "=COUNTA(Data!G3:G12)"),
]

for r, row in enumerate(lookup_rows, start=1):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
        if r == 1:
            ws.cell(row=r, column=c).font = Font(bold=True)
            ws.cell(row=r, column=c).fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )
        ws.cell(row=r, column=c).border = thin_border

# -----------------------------
# Final touches & Save
# -----------------------------
# Set some default fonts/alignments for headers already done. Adjust row heights lightly.
for wsname in ["Data", "Tasks", "Answers", "Checklist", "Lookup"]:
    wss = wb[wsname]
    for row in wss.iter_rows(
        min_row=1, max_row=wss.max_row, min_col=1, max_col=wss.max_column
    ):
        for cell in row:
            if cell.row in (1, 2) and isinstance(cell.value, str) and cell.value:
                cell.alignment = Alignment(vertical="center")

# Default active sheet on open
wb.active = wb["Instructions"]

wb.save("Core_Functions_Practice.xlsx")
print("Workbook created: Core_Functions_Practice.xlsx")
