from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference


# ---------- Helper styling ----------
def set_col_width(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def header_style(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DCE6F1")
    cell.alignment = Alignment(vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def box(ws, cell_range):
    thin = Side(style="thin")
    for row in ws[cell_range]:
        for c in row:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------- Build workbook ----------
wb = Workbook()

# Sheet: Instructions
wsI = wb.active
wsI.title = "Instructions"
wsI["A1"] = "N Level Excel: Conditional Counting (COUNTIF / COUNTIFS)"
wsI["A1"].font = Font(size=14, bold=True)
lines = [
    "Goal: Practice counting with conditions using COUNTIF (one condition) and COUNTIFS (multiple conditions).",
    "",
    "How to use this workbook:",
    "1) Go to the Data sheet to view the data table.",
    "2) Open the Tasks sheet and write your formulas in the Answer cells (column C).",
    "3) Watch the Answer Check column turn Green (Correct) when your formula matches the expected value.",
    "4) Use the Hints sheet if you get stuck; check final solutions in the Answers sheet.",
    "",
    "Keyboard tips (Windows): Enter formula =, confirm with Enter, copy with Ctrl+C, paste with Ctrl+V, fill down with Ctrl+D.",
    "Mac tips: Cmd instead of Ctrl.",
    "",
    "Learning focus today:",
    '- COUNTIF(range, criteria)   e.g. =COUNTIF(C2:C41, ">100")',
    '- COUNTIFS(range1, crit1, range2, crit2, ...)   e.g. =COUNTIFS(B2:B41, "Singapore", C2:C41, ">100")',
]
for r, text in enumerate(lines, start=3):
    wsI[f"A{r}"] = text
set_col_width(wsI, {"A": 110})

# Sheet: Data
wsD = wb.create_sheet("Data")
headers = ["Name", "Country", "Sales", "Channel"]
wsD.append(headers)
data = [
    # Name, Country, Sales, Channel
    ["Alex", "Singapore", 120, "Online"],
    ["Ben", "Malaysia", 80, "Store"],
    ["Clara", "Singapore", 60, "Online"],
    ["Devi", "Singapore", 200, "Store"],
    ["Ethan", "Malaysia", 150, "Online"],
    ["Farah", "Indonesia", 95, "Store"],
    ["Gwen", "Singapore", 130, "Online"],
    ["Hadi", "Indonesia", 40, "Store"],
    ["Iris", "Malaysia", 220, "Online"],
    ["Jamal", "Singapore", 55, "Store"],
    ["Kara", "Indonesia", 175, "Online"],
    ["Leo", "Malaysia", 45, "Online"],
    ["Maya", "Singapore", 160, "Store"],
    ["Nora", "Indonesia", 85, "Online"],
    ["Omar", "Malaysia", 110, "Store"],
    ["Pia", "Singapore", 30, "Online"],
    ["Qadir", "Indonesia", 125, "Store"],
    ["Rina", "Malaysia", 155, "Online"],
    ["Sam", "Singapore", 99, "Store"],
    ["Tara", "Indonesia", 60, "Online"],
    ["Uma", "Malaysia", 140, "Store"],
    ["Vik", "Singapore", 175, "Online"],
    ["Wes", "Indonesia", 210, "Store"],
    ["Xena", "Malaysia", 70, "Online"],
    ["Yasmin", "Singapore", 115, "Store"],
    ["Zack", "Indonesia", 135, "Online"],
    ["Ari", "Malaysia", 190, "Store"],
    ["Brynn", "Singapore", 52, "Online"],
    ["Cody", "Indonesia", 48, "Store"],
    ["Dina", "Malaysia", 101, "Online"],
    ["Eli", "Singapore", 149, "Store"],
    ["Fio", "Indonesia", 151, "Online"],
    ["Gabe", "Malaysia", 67, "Store"],
    ["Hana", "Singapore", 88, "Online"],
    ["Ivan", "Indonesia", 112, "Store"],
    ["Jae", "Malaysia", 158, "Online"],
    ["Kimi", "Singapore", 200, "Store"],
    ["Lia", "Indonesia", 77, "Online"],
    ["Milo", "Malaysia", 89, "Store"],
    ["Nia", "Singapore", 105, "Online"],
]
for row in data:
    wsD.append(row)

# Style header
for c in wsD[1]:
    header_style(c)

# Table formatting
end_row = wsD.max_row
tbl = Table(displayName="SalesTbl", ref=f"A1:D{end_row}")
style = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
)
tbl.tableStyleInfo = style
wsD.add_table(tbl)

# widen columns
set_col_width(wsD, {"A": 14, "B": 14, "C": 10, "D": 12})

# Summary (for chart): counts by Country using COUNTIF
wsD["F1"] = "Summary: Count by Country"
wsD["F1"].font = Font(bold=True)
wsD["F3"] = "Country"
wsD["G3"] = "Count"
summary_countries = ["Singapore", "Malaysia", "Indonesia"]
for i, ctry in enumerate(summary_countries, start=4):
    wsD[f"F{i}"] = ctry
    wsD[f"G{i}"] = f'=COUNTIF(B2:B{end_row},"{ctry}")'
header_style(wsD["F3"])
header_style(wsD["G3"])
box(wsD, f"F3:G{4 + len(summary_countries) - 1}")

# Chart
chart = BarChart()
chart.title = "Counts by Country"
chart.y_axis.title = "Count"
chart.x_axis.title = "Country"
data_ref = Reference(wsD, min_col=7, min_row=3, max_row=3 + len(summary_countries))
cats_ref = Reference(wsD, min_col=6, min_row=4, max_row=3 + len(summary_countries))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
wsD.add_chart(chart, "I3")

# Sheet: Lookup (for dropdowns)
wsL = wb.create_sheet("Lookup")
wsL.append(["Countries"])
for c in summary_countries:
    wsL.append([c])
wsL["D1"] = "Channels"
for i, ch in enumerate(["Online", "Store"], start=2):
    wsL.cell(row=i, column=4, value=ch)
set_col_width(wsL, {"A": 18, "D": 18})

# Sheet: Tasks
wsT = wb.create_sheet("Tasks")
set_col_width(wsT, {"A": 60, "B": 18, "C": 18, "D": 18})
wsT["A1"] = "Tasks: Enter your COUNTIF / COUNTIFS formulas in column C (Answer)."
wsT["A1"].font = Font(size=12, bold=True)

task_rows = [
    ("1) Count how many sales are LESS than 100.", "Number", ""),
    ("2) Count how many Malaysia sales are GREATER than 100.", "Number", ""),
    (
        "3) Count how many Singapore sales are BETWEEN 50 and 150 (inclusive).",
        "Number",
        "",
    ),
    ("4) Count how many Online sales are from Indonesia.", "Number", ""),
    ("5) Count how many names start with the letter A.", "Number", ""),
]
wsT.append(["Task", "Expected Type", "Answer (your formula result)", "Answer Check"])
for c in wsT[2]:
    header_style(c)

start_r = 3
for i, (t, ttype, _) in enumerate(task_rows, start=start_r):
    wsT[f"A{i}"] = t
    wsT[f"B{i}"] = ttype
    # C = student input cell (result of their formula)
    # D = checker comparing to Answers sheet
    wsT[f"D{i}"] = f'=IF(C{i}=Answers!B{i - (start_r - 3) + 2},"Correct","Check again")'

# Data validation dropdowns (optional helpers)
dv_country = DataValidation(type="list", formula1="=Lookup!$A$2:$A$4", allow_blank=True)
dv_channel = DataValidation(type="list", formula1="=Lookup!$D$2:$D$3", allow_blank=True)
wsT.add_data_validation(dv_country)
wsT.add_data_validation(dv_channel)
# Place helper dropdown cells for student experimentation
wsT["A10"] = "Helper dropdowns (optional for your own tests):"
wsT["B11"] = "Country:"
wsT["C11"] = ""
wsT["B12"] = "Channel:"
wsT["C12"] = ""
dv_country.add(wsT["C11"])
dv_channel.add(wsT["C12"])
box(wsT, "A10:D12")

# Conditional formatting for Answer Check
from openpyxl.formatting.rule import FormulaRule

wsT.conditional_formatting.add(
    f"D{start_r}:D{start_r + len(task_rows) - 1}",
    FormulaRule(
        formula=[f'INDIRECT("D"&ROW())="Correct"'],
        stopIfTrue=True,
        fill=PatternFill("solid", fgColor="C6EFCE"),
    ),
)
wsT.conditional_formatting.add(
    f"D{start_r}:D{start_r + len(task_rows) - 1}",
    FormulaRule(
        formula=[f'INDIRECT("D"&ROW())="Check again"'],
        stopIfTrue=True,
        fill=PatternFill("solid", fgColor="FFC7CE"),
    ),
)

# Sheet: Hints
wsH = wb.create_sheet("Hints")
set_col_width(wsH, {"A": 110})
hints = [
    "General tips:",
    "- COUNTIF uses ONE condition: =COUNTIF(range, crit)",
    "- COUNTIFS uses MULTIPLE conditions: =COUNTIFS(rng1, crit1, rng2, crit2, ...)",
    '- Put text and comparison operators in quotes, e.g. "Singapore", ">100".',
    "",
    "Task hints:",
    '1) Use COUNTIF on Sales column C: criteria is "<100".',
    '2) Use COUNTIFS with Country (B) and Sales (C): ">100".',
    '3) Use COUNTIFS with two Sales conditions: ">=50" and "<=150" and Country = "Singapore".',
    "4) Use COUNTIFS with Channel (D) and Country (B).",
    '5) Use COUNTIF on Names (A) with a wildcard pattern: "A*".',
]
for r, t in enumerate(hints, start=1):
    wsH[f"A{r}"] = t

# Sheet: Answers
wsA = wb.create_sheet("Answers")
set_col_width(wsA, {"A": 60, "B": 18, "C": 90})
wsA.append(["Task", "Correct Result", "Suggested Formula"])
for c in wsA[1]:
    header_style(c)

# Calculate end_row dynamically for formulas
last = wsD.max_row
answers = [
    (
        "1) Count sales < 100",
        f'=COUNTIF(Data!C2:C{last}, "<100")',
        f'=COUNTIF(Data!C2:C{last}, "<100")',
    ),
    (
        "2) Malaysia sales > 100",
        f'=COUNTIFS(Data!B2:B{last}, "Malaysia", Data!C2:C{last}, ">100")',
        f'=COUNTIFS(Data!B2:B{last}, "Malaysia", Data!C2:C{last}, ">100")',
    ),
    (
        "3) Singapore sales between 50 and 150 (inclusive)",
        f'=COUNTIFS(Data!B2:B{last}, "Singapore", Data!C2:C{last}, ">=50", Data!C2:C{last}, "<=150")',
        f'=COUNTIFS(Data!B2:B{last}, "Singapore", Data!C2:C{last}, ">=50", Data!C2:C{last}, "<=150")',
    ),
    (
        "4) Online sales from Indonesia",
        f'=COUNTIFS(Data!D2:D{last}, "Online", Data!B2:B{last}, "Indonesia")',
        f'=COUNTIFS(Data!D2:D{last}, "Online", Data!B2:B{last}, "Indonesia")',
    ),
    (
        "5) Names starting with A",
        f'=COUNTIF(Data!A2:A{last}, "A*")',
        f'=COUNTIF(Data!A2:A{last}, "A*")',
    ),
]
for row in answers:
    wsA.append(row)

# Sheet: Checklist
wsC = wb.create_sheet("Checklist")
set_col_width(wsC, {"A": 90, "B": 14})
wsC.append(["Skill", "Done?"])
for c in wsC[1]:
    header_style(c)
skills = [
    "I can use COUNTIF for a single condition.",
    "I can use COUNTIFS for multiple conditions.",
    'I know to put text and operators in quotes ("Singapore", ">100").',
    "I ensure COUNTIFS ranges are the same size.",
    "I can use wildcards like A* for text patterns.",
]
for s in skills:
    wsC.append([s, ""])
box(wsC, f"A1:B{wsC.max_row}")

# Neaten up Tasks header row
for cell in wsT[2]:
    cell.alignment = Alignment(vertical="center")

# Freeze panes for usability
wsD.freeze_panes = "A2"
wsT.freeze_panes = "A3"
wsA.freeze_panes = "A2"

# Final save
wb.save("NLevel_COUNTIFS_Practice.xlsx")
print("Workbook created: NLevel_COUNTIFS_Practice.xlsx")
