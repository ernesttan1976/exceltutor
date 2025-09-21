# make_charts_practice.py
# Creates an N Level Excel starter workbook focused on Charts & Visuals.
# Sheets: Instructions, Data, Tasks, Hints, Answers, Checklist, Lookup, Charts

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation


# ---------- Helper formatting ----------
def title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(size=14, bold=True)
    ws.merge_cells(
        start_row=ws[cell].row,
        start_column=ws[cell].column,
        end_row=ws[cell].row,
        end_column=ws.max_column
        if ws.max_column > ws[cell].column
        else ws[cell].column,
    )
    ws[cell].alignment = Alignment(horizontal="left", vertical="center")


def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def header_row(ws, row):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ---------- Build workbook ----------
wb = Workbook()

# 1) Instructions
ws = wb.active
ws.title = "Instructions"
set_col_widths(ws, {"A": 80})
instructions = [
    "N Level Excel — Charts & Visuals Starter",
    "",
    "What’s inside:",
    "• Data: Sample monthly sales + product share.",
    "• Tasks: Step-by-step practice (Column, Line, Pie).",
    "• Hints: Formula and chart tips.",
    "• Answers: Suggested answers and example formulas.",
    "• Checklist: Self-check before submitting work.",
    "• Lookup: Reference of common functions.",
    "• Charts: Pre-built Column, Line, and Pie charts.",
    "",
    "How to use:",
    "1) Read the Tasks sheet and follow each step.",
    "2) Use Hints if you’re stuck; check Answers when done.",
    "3) Edit values on the Data sheet and watch charts update.",
    "4) Practice formatting titles, axis labels, and data labels.",
    "",
    "Keyboard shortcuts (Win / Mac):",
    "• Select entire column: Ctrl+Space / Cmd+Space",
    "• Select entire row: Shift+Space / Shift+Space",
    "• Insert chart quickly: Alt+N then choose chart / Ribbon",
]
for i, line in enumerate(instructions, start=1):
    ws[f"A{i}"] = line
for row in ws.iter_rows(min_row=1, max_row=len(instructions), min_col=1, max_col=1):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# 2) Data
ws = wb.create_sheet("Data")
set_col_widths(ws, {"A": 12, "B": 10, "C": 10, "D": 12, "F": 16, "G": 10})
ws["A1"] = "Month"
ws["B1"] = "Sales"
ws["C1"] = "Budget"
ws["D1"] = "Category"
header_row(ws, 1)
months = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]
sales = [200, 300, 250, 400, 350, 380, 360, 420, 390, 410, 370, 430]
budget = [220, 280, 260, 380, 360, 370, 350, 400, 400, 405, 380, 420]
cats = ["A", "A", "A", "A", "B", "B", "B", "B", "C", "C", "C", "C"]
for r, (m, s, b, c) in enumerate(zip(months, sales, budget, cats), start=2):
    ws[f"A{r}"] = m
    ws[f"B{r}"] = s
    ws[f"C{r}"] = b
    ws[f"D{r}"] = c

# Add a table for the monthly data
table = Table(displayName="tblMonthly", ref=f"A1:D{1 + len(months)}")
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
table.tableStyleInfo = style
ws.add_table(table)

# Second dataset for Pie (product share)
ws["F1"] = "Product"
ws["G1"] = "Units"
header_row(ws, 1)
products = [("Alpha", 120), ("Bravo", 80), ("Charlie", 60), ("Delta", 40)]
for i, (p, u) in enumerate(products, start=2):
    ws[f"F{i}"] = p
    ws[f"G{i}"] = u

pie_table = Table(displayName="tblProducts", ref="F1:G5")
pie_style = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
pie_table.tableStyleInfo = pie_style
ws.add_table(pie_table)

# Helpful summary cells
ws["I1"] = "Quick Stats"
ws["I1"].font = Font(bold=True)
ws["I2"] = "Total Sales"
ws["J2"] = f"=SUM(B2:B{1 + len(months)})"
ws["I3"] = "Average Sales"
ws["J3"] = f"=AVERAGE(B2:B{1 + len(months)})"
ws["I4"] = "Max Month"
ws["J4"] = (
    f"=XLOOKUP(MAX(B2:B{1 + len(months)}),B2:B{1 + len(months)},A2:A{1 + len(months)})"
)
ws["I5"] = "Min Month"
ws["J5"] = (
    f"=XLOOKUP(MIN(B2:B{1 + len(months)}),B2:B{1 + len(months)},A2:A{1 + len(months)})"
)

# Data validation example (drop-down for category)
dv = DataValidation(type="list", formula1='"A,B,C"', allow_blank=True)
ws.add_data_validation(dv)
dv.add("D2:D13")

# 3) Tasks
ws = wb.create_sheet("Tasks")
set_col_widths(ws, {"A": 70, "B": 30})
tasks = [
    (
        "Starter (Column)",
        "On Data sheet, select A1:B13, Insert → Column → Clustered Column. Add chart title 'Monthly Sales'. Add data labels.",
    ),
    (
        "Core (Line)",
        "Create a line chart showing Sales vs Month. Add axis titles: Month (X), Sales (Y). Add a legend.",
    ),
    (
        "Core (Compare)",
        "Create a column chart comparing Sales and Budget (A1:C13). Use a meaningful title and show data labels.",
    ),
    (
        "Stretch (Pie)",
        "Build a pie chart from Product/Units (F1:G5). Show percentages and a clear title.",
    ),
    (
        "Stretch (Format)",
        "Change chart colors, bold the title, and adjust the chart area so labels are readable.",
    ),
    (
        "Challenge",
        "Which month exceeded Budget by the largest margin? Compute a helper column 'Variance' = Sales - Budget and label the max with conditional formatting.",
    ),
]
ws["A1"] = "Task"
ws["B1"] = "Notes / Check"
header_row(ws, 1)
for i, (t, n) in enumerate(tasks, start=2):
    ws[f"A{i}"] = t
    ws[f"B{i}"] = n
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
    for cell in row:
        cell.border = thin_border

# 4) Hints
ws = wb.create_sheet("Hints")
set_col_widths(ws, {"A": 80, "B": 60})
ws.append(["Topic", "Hint"])
header_row(ws, 1)
ws.append(
    [
        "Selecting data",
        "Include headers (Month, Sales) so Excel builds a clean legend/axis.",
    ]
)
ws.append(
    [
        "Data labels",
        "After inserting a chart, use the + button (Chart Elements) → Data Labels.",
    ]
)
ws.append(["Axis titles", "Use + button → Axis Titles. Name X: Month, Y: Sales."])
ws.append(
    ["Helper column", "In Data!E1 type 'Variance', in E2 enter =B2-C2 and fill down."]
)
ws.append(
    ["Find max variance", "Use =MAX(E2:E13) to get the largest positive variance."]
)
ws.append(
    [
        "Month of max variance",
        "Use =XLOOKUP(MAX(E2:E13),E2:E13,A2:A13) to return the month.",
    ]
)

# 5) Answers (suggested)
ws = wb.create_sheet("Answers")
set_col_widths(ws, {"A": 34, "B": 60})
ws["A1"] = "Question"
ws["B1"] = "Answer (Example)"
header_row(ws, 1)
answers = [
    ("Variance formula", "=B2-C2 (fill down)"),
    ("Largest positive variance", "=MAX(Data!E2:E13)"),
    (
        "Month with largest variance",
        "=XLOOKUP(MAX(Data!E2:E13),Data!E2:E13,Data!A2:A13)",
    ),
    ("Total Sales", "=SUM(Data!B2:B13)"),
    ("Average Sales", "=AVERAGE(Data!B2:B13)"),
]
for i, (q, a) in enumerate(answers, start=2):
    ws[f"A{i}"] = q
    ws[f"B{i}"] = a

# 6) Checklist
ws = wb.create_sheet("Checklist")
set_col_widths(ws, {"A": 60, "B": 14})
ws.append(["Item", "Done?"])
header_row(ws, 1)
check_items = [
    "Chart has a clear, descriptive title",
    "Axes are labeled (where relevant)",
    "Appropriate chart type chosen",
    "Data labels added (where useful)",
    "Legend is clear / not cluttered",
    "Numbers formatted correctly",
    "No overlapping labels",
    "Colors improve readability",
]
for item in check_items:
    ws.append([item, "Yes/No"])

# 7) Lookup (quick reference)
ws = wb.create_sheet("Lookup")
set_col_widths(ws, {"A": 24, "B": 80})
ws.append(["Function", "Usage"])
header_row(ws, 1)
lookups = [
    ("SUM", "Add numbers: =SUM(B2:B13)"),
    ("AVERAGE", "Mean: =AVERAGE(B2:B13)"),
    ("MIN / MAX", "Smallest / Largest: =MIN(B2:B13), =MAX(B2:B13)"),
    ("COUNT / COUNTA", "Count numbers / non-blanks"),
    ("IF", 'Basic decision: =IF(B2>=C2,"Above","Below")'),
    ("COUNTIF", 'Count matching: =COUNTIF(D2:D13,"A")'),
    ("XLOOKUP", "Find a value: =XLOOKUP(lookup, lookup_range, return_range)"),
    ("TEXTJOIN", 'Combine text: =TEXTJOIN(", ",TRUE,A2:A5)'),
]
for f, u in lookups:
    ws.append([f, u])

# 8) Charts (pre-built)
ws = wb.create_sheet("Charts")
set_col_widths(ws, {"A": 16, "B": 16, "C": 16, "D": 16})

# References to Data sheet
data_ws = wb["Data"]
# Column/Line ranges
cat_ref = Reference(
    data_ws, min_col=1, min_row=1, max_row=13
)  # A1:A13 (Month header + months)
sales_ref = Reference(data_ws, min_col=2, min_row=1, max_row=13)  # B1:B13
budget_ref = Reference(data_ws, min_col=3, min_row=1, max_row=13)  # C1:C13

# Column Chart (Sales vs Month)
bar = BarChart()
bar.type = "col"
bar.title = "Monthly Sales (Column)"
bar.y_axis.title = "Sales"
bar.x_axis.title = "Month"
bar.add_data(sales_ref, titles_from_data=True)
bar.set_categories(cat_ref)
bar.dataLabels = DataLabelList()
bar.dataLabels.showVal = True
ws.add_chart(bar, "A2")

# Line Chart (Sales & Budget vs Month)
line = LineChart()
line.title = "Sales vs Budget (Line)"
line.y_axis.title = "Value"
line.x_axis.title = "Month"
line.add_data(sales_ref, titles_from_data=True)
line.add_data(budget_ref, titles_from_data=True)
line.set_categories(cat_ref)
ws.add_chart(line, "J2")

# Pie Chart (Product Share)
prod_cat = Reference(data_ws, min_col=6, min_row=2, max_row=5)  # F2:F5 products
units_ref = Reference(
    data_ws, min_col=7, min_row=1, max_row=5
)  # G1:G5 (header + units)
pie = PieChart()
pie.title = "Product Share (Pie)"
pie.add_data(units_ref, titles_from_data=True)
pie.set_categories(prod_cat)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showLeaderLines = True
ws.add_chart(pie, "A20")

# Cosmetic: small headers on Charts
ws["A1"] = "Pre-built Charts"
ws["A1"].font = Font(size=12, bold=True)

# Save
wb.save("Charts_Practice.xlsx")
print("Created Charts_Practice.xlsx")
