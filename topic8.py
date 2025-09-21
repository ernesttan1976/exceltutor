# make_text_functions_workbook.py
# Creates "Text_Functions_Practice.xlsx" with sheets:
# Instructions, Data, Tasks, Hints, Answers, Checklist, Lookup
# Includes sample data, formulas (LEFT, RIGHT, MID, LEN, CONCAT, TEXTJOIN),
# an example chart, basic table formatting, and dropdown validation.

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ---------- helpers ----------
def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def header_style(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def title(ws, text, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=14, bold=True)
    c.alignment = Alignment(horizontal="left")


def add_table(ws, ref, name):
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


# ---------- workbook ----------
wb = Workbook()

# rename default sheet to Instructions
ws_instr = wb.active
ws_instr.title = "Instructions"

# ---------- Instructions ----------
title(ws_instr, "Text Functions Practice – Instructions", row=1)
instr_lines = [
    "Goal: Practice LEFT, RIGHT, MID, LEN, CONCAT, and TEXTJOIN to clean and combine text.",
    "",
    "How to use this workbook:",
    "1) Read Hints and Lookup for function syntax and examples.",
    "2) Open Tasks: complete the yellow cells ONLY (enter formulas).",
    "3) Use data from the Data sheet when a task references it.",
    "4) Check your work on the Answers sheet (formulas are shown).",
    "5) Use the Checklist to track what you’ve mastered.",
    "",
    "Tip: Spaces count as characters in LEN. If you see unexpected counts, check for spaces!",
    "Shortcuts (Windows): Enter formula =, confirm with Enter; copy down: Ctrl+D; fill right: Ctrl+R.",
    "Mac: copy down ⌘+D; fill right ⌘+R.",
]
for i, line in enumerate(instr_lines, start=3):
    ws_instr.cell(row=i, column=1, value=line)
set_col_widths(ws_instr, {"A": 110})

# ---------- Data ----------
ws_data = wb.create_sheet("Data")
title(ws_data, "Sample Data", row=1)
headers = ["ID", "Full Name", "Product Code", "Item A", "Item B", "Item C", "City"]
data_rows = [
    [101, "Lim Wei Ming", "INV2025-AB", "Apple", "Mango", "Pear", "Singapore"],
    [102, "Tan Siew Ling", "INV2024-ZX", "Orange", "Kiwi", "Banana", "Johor Bahru"],
    [103, "Nur Aisyah", "ORD2030-Q1", "Grape", "", "Melon", "Kuala Lumpur"],
    [104, "Goh Jun Hao", "REF2022-PQ", "Pear", "Apple", "", "Singapore"],
    [105, "Chong Zi Xuan", "INV2025-CD", "", "Lychee", "Longan", "Malacca"],
]
ws_data.append(headers)
for r in data_rows:
    ws_data.append(r)

# style header
for col in range(1, len(headers) + 1):
    header_style(ws_data.cell(row=2, column=col))

# table
last_row = 2 + len(data_rows)
last_col = len(headers)
ref = f"A2:{get_column_letter(last_col)}{last_row}"
add_table(ws_data, ref, "tblData")

set_col_widths(ws_data, {"A": 8, "B": 20, "C": 15, "D": 12, "E": 12, "F": 12, "G": 18})

# ---------- Tasks ----------
ws_tasks = wb.create_sheet("Tasks")
title(ws_tasks, "Tasks – Enter formulas in yellow cells only", row=1)
task_headers = [
    "Task #",
    "Description",
    "Input / Reference",
    "Your Formula",
    "Expected Result (auto-check)",
]
ws_tasks.append(task_headers)
for c in range(1, len(task_headers) + 1):
    header_style(ws_tasks.cell(row=2, column=c))

yellow = PatternFill(start_color="FFFDE599", end_color="FFFDE599", fill_type="solid")

tasks = [
    (1, "LEN of a phrase", 'Text: "Excel Skills"', "", 'LEN("Excel Skills")'),
    (2, "First 4 letters (LEFT)", 'Text: "Singapore"', "", 'LEFT("Singapore",4)'),
    (
        3,
        "Extract year using MID",
        'From Data!C3 e.g. "INV2024-ZX"',
        "",
        "MID(Data!C3,4,4)",
    ),
    (
        4,
        "Join two words with space (CONCAT)",
        'Words: "N Level" + "Excel"',
        "",
        'CONCAT("N Level"," ","Excel")',
    ),
    (
        5,
        "Extract middle name (MID)",
        'From Data!B3 "Tan Siew Ling"',
        "",
        'MID(Data!B3,5,4)   -> "Siew"',
    ),
    (
        6,
        "TEXTJOIN with commas, ignore blanks",
        "Data!D2:F2",
        "",
        'TEXTJOIN(", ",TRUE,Data!D2:F2)',
    ),
    (
        7,
        "Build short code: SURNAME(3)-LASTNAME(4)",
        "From Data!B2",
        "",
        'CONCAT(LEFT(Data!B2,3),"-",RIGHT(Data!B2,4))',
    ),
    (
        8,
        "LEN of Full Name (including spaces)",
        "From Data!B2:B6",
        "",
        "LEN(Data!B2) etc.",
    ),
]

start_row = 3
for r_idx, (num, desc, input_ref, placeholder, check) in enumerate(
    tasks, start=start_row
):
    ws_tasks.cell(row=r_idx, column=1, value=num)
    ws_tasks.cell(row=r_idx, column=2, value=desc)
    ws_tasks.cell(row=r_idx, column=3, value=input_ref)
    fcell = ws_tasks.cell(row=r_idx, column=4, value="")  # user formula goes here
    fcell.fill = yellow
    ws_tasks.cell(row=r_idx, column=5, value=check)

set_col_widths(ws_tasks, {"A": 8, "B": 38, "C": 32, "D": 40, "E": 40})

# simple dropdown to choose delimiter for TEXTJOIN (optional use in Tasks #6)
ws_tasks.cell(row=12, column=1, value="Options")
ws_tasks.cell(row=13, column=1, value="Delimiter Choice")
ws_tasks.cell(row=13, column=2, value=", ")
ws_tasks.cell(row=14, column=2, value="; ")
ws_tasks.cell(row=15, column=2, value=" | ")

dv = DataValidation(type="list", formula1="=$B$14:$B$15", allow_blank=True)
ws_tasks.add_data_validation(dv)
dv.add(ws_tasks["B13"])

# ---------- Hints ----------
ws_hints = wb.create_sheet("Hints")
title(ws_hints, "Hints – Syntax & Tips", row=1)
hints = [
    ["Function", "Syntax", "What it does", "Example"],
    [
        "LEFT",
        "LEFT(text, num_chars)",
        "Takes characters from the left",
        'LEFT("Singapore",3) -> "Sin"',
    ],
    [
        "RIGHT",
        "RIGHT(text, num_chars)",
        "Takes characters from the right",
        'RIGHT("Singapore",4) -> "pore"',
    ],
    [
        "MID",
        "MID(text, start_num, num_chars)",
        "Takes characters from the middle",
        'MID("Singapore",4,3) -> "gap"',
    ],
    ["LEN", "LEN(text)", "Counts characters incl. spaces", 'LEN("Excel Skills") -> 12'],
    [
        "CONCAT",
        "CONCAT(text1, [text2], ...)",
        "Joins text items",
        'CONCAT("N Level"," ","Excel") -> "N Level Excel"',
    ],
    [
        "TEXTJOIN",
        "TEXTJOIN(delimiter, ignore_empty, text1, ...)",
        "Joins ranges with a delimiter",
        'TEXTJOIN(", ",TRUE,Data!D2:F2)',
    ],
    [
        "Tip",
        "",
        "Spaces count! Use TRIM(text) if there are stray spaces.",
        'TRIM("  hello ") -> "hello"',
    ],
]
for row in hints:
    ws_hints.append(row)
for c in range(1, 5):
    header_style(ws_hints.cell(row=2, column=c))
add_table(ws_hints, "A2:D9", "tblHints")
set_col_widths(ws_hints, {"A": 14, "B": 38, "C": 42, "D": 46})

# ---------- Answers ----------
ws_ans = wb.create_sheet("Answers")
title(ws_ans, "Answers – Completed formulas", row=1)

ans_headers = [
    "Row",
    "Full Name",
    "LEN",
    "First 3 (LEFT)",
    "Last 4 (RIGHT)",
    "Year (MID)",
    "Short Code",
    "Items (TEXTJOIN)",
]
ws_ans.append(ans_headers)
for c in range(1, len(ans_headers) + 1):
    header_style(ws_ans.cell(row=2, column=c))

# Fill formulas for each row in Data
# Data rows start at Data!2 to Data!6
ans_row_start = 3
for i in range(2, 7):
    target_row = ans_row_start + (i - 2)
    ws_ans.cell(row=target_row, column=1, value=i - 1)  # Row #
    ws_ans.cell(row=target_row, column=2, value=f"=Data!B{i}")
    ws_ans.cell(row=target_row, column=3, value=f"=LEN(Data!B{i})")
    ws_ans.cell(row=target_row, column=4, value=f"=LEFT(Data!B{i},3)")
    ws_ans.cell(row=target_row, column=5, value=f"=RIGHT(Data!B{i},4)")
    # If code like INV2025-AB, year is chars 4-7
    ws_ans.cell(row=target_row, column=6, value=f"=MID(Data!C{i},4,4)")
    ws_ans.cell(
        row=target_row,
        column=7,
        value=f'=CONCAT(LEFT(Data!B{i},3),"-",RIGHT(Data!B{i},4))',
    )
    ws_ans.cell(row=target_row, column=8, value=f'=TEXTJOIN(", ",TRUE,Data!D{i}:F{i})')

add_table(ws_ans, f"A2:H{ans_row_start + 5}", "tblAnswers")
set_col_widths(
    ws_ans, {"A": 6, "B": 22, "C": 8, "D": 16, "E": 16, "F": 12, "G": 16, "H": 28}
)

# Chart: bar chart of name lengths
chart = BarChart()
chart.title = "Full Name Character Count"
chart.y_axis.title = "Characters"
chart.x_axis.title = "Row"

data_ref = Reference(
    ws_ans, min_col=3, min_row=2, max_row=ans_row_start + 5
)  # LEN column incl header
cats_ref = Reference(
    ws_ans, min_col=1, min_row=3, max_row=ans_row_start + 5
)  # Row numbers
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws_ans.add_chart(chart, "J3")

# ---------- Checklist ----------
ws_check = wb.create_sheet("Checklist")
title(ws_check, "Checklist – Tick off when done", row=1)
check_items = [
    ["Skill", "Done? (Y/N)", "Notes"],
    ["Use LEN to count characters", "", ""],
    ["Extract with LEFT and RIGHT", "", ""],
    ["Extract with MID (middle)", "", ""],
    ["Join with CONCAT", "", ""],
    ["Join a range with TEXTJOIN, ignore blanks", "", ""],
    ["Understand that spaces count in LEN", "", ""],
]
for row in check_items:
    ws_check.append(row)
for c in range(1, 4):
    header_style(ws_check.cell(row=2, column=c))
add_table(ws_check, "A2:C8", "tblChecklist")
set_col_widths(ws_check, {"A": 40, "B": 14, "C": 46})

# ---------- Lookup ----------
ws_lookup = wb.create_sheet("Lookup")
title(ws_lookup, "Quick Reference – Text Functions", row=1)
lookup_rows = [
    ["Function", "Key Arguments", "Notes / Example"],
    ["LEFT", "text, num_chars", 'e.g. LEFT("Hello",2) -> "He"'],
    ["RIGHT", "text, num_chars", 'e.g. RIGHT("Hello",3) -> "llo"'],
    [
        "MID",
        "text, start_num, num_chars",
        'e.g. MID("Hello",2,2) -> "el" (starts at H=1, e=2)',
    ],
    ["LEN", "text", "Counts spaces too."],
    ["CONCAT", "text1, [text2]…", "Simple join, no delimiter built-in."],
    [
        "TEXTJOIN",
        "delimiter, ignore_empty, text1…",
        'TEXTJOIN(", ",TRUE,range) joins with commas and skips blanks.',
    ],
    ["TRIM", "text", "Removes extra spaces (handy before LEN)."],
]
for row in lookup_rows:
    ws_lookup.append(row)
for c in range(1, 4):
    header_style(ws_lookup.cell(row=2, column=c))
add_table(ws_lookup, "A2:C9", "tblLookup")
set_col_widths(ws_lookup, {"A": 14, "B": 28, "C": 70})

# Freeze panes & nice view settings
ws_tasks.freeze_panes = "A3"
ws_ans.freeze_panes = "A3"
ws_data.freeze_panes = "A3"

# Save
filename = "Text_Functions_Practice.xlsx"
wb.save(filename)

print(f"Workbook created: {filename}")
