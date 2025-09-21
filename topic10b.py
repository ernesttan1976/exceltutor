# build_conditional_formatting_workbook.py
# Creates an Excel practice workbook for Conditional Formatting (N Level)
# Sheets: Instructions, Data, Tasks, Hints, Answers, Checklist, Lookup

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, DataBarRule, Rule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference


# ---------- Helpers ----------
def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def title(ws, text, cell="A1"):
    ws[cell] = text
    ws[cell].font = Font(size=16, bold=True)
    return ws[cell]


def add_table(ws, ref, name, style="TableStyleMedium9"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)


thin = Side(style="thin", color="CCCCCC")
thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)

# ---------- Workbook ----------
wb = Workbook()

# Remove default sheet and create named ones in desired order
wb.remove(wb.active)
ws_instr = wb.create_sheet("Instructions")
ws_data = wb.create_sheet("Data")
ws_tasks = wb.create_sheet("Tasks")
ws_hints = wb.create_sheet("Hints")
ws_answers = wb.create_sheet("Answers")
ws_check = wb.create_sheet("Checklist")
ws_lookup = wb.create_sheet("Lookup")

# ---------- Instructions ----------
title(ws_instr, "Excel Practice: Conditional Formatting")
ws_instr["A3"] = "Objective"
ws_instr["A3"].font = Font(bold=True)
ws_instr["B3"] = (
    "Highlight cells automatically based on rules (e.g., marks below 50 turn red)."
)

ws_instr["A5"] = "Why it matters"
ws_instr["A5"].font = Font(bold=True)
ws_instr["B5"] = (
    "Makes tables easier to read, spots top/bottom values, and saves time versus manual checking."
)

ws_instr["A7"] = "Steps"
ws_instr["A7"].font = Font(bold=True)
steps = [
    "Select the range you want to format.",
    "Home → Conditional Formatting → choose a rule (Less Than, Greater Than, Between, Top/Bottom, Data Bars...).",
    "Enter the condition (e.g., 50). Pick a format (e.g., red fill). Click OK.",
]
for i, s in enumerate(steps, start=8):
    ws_instr[f"B{i}"] = f"{i - 7}. {s}"

ws_instr["A12"] = "Worked example"
ws_instr["A12"].font = Font(bold=True)
ws_instr["B12"] = "On the Data sheet, highlight Marks < 50 in red."

ws_instr["A14"] = "How to use this file"
ws_instr["A14"].font = Font(bold=True)
howto = [
    "Go to the Data sheet and review the table.",
    "Open the Tasks sheet and complete each task in order.",
    "Use Hints if stuck; check visual results against the Answers sheet.",
    "Tick off items in the Checklist when done.",
]
for i, s in enumerate(howto, start=15):
    ws_instr[f"B{i}"] = f"- {s}"

set_col_widths(ws_instr, {"A": 18, "B": 90})
ws_instr.freeze_panes = "A8"

# ---------- Data ----------
title(ws_data, "Student Marks")
headers = ["Name", "Class", "Marks", "Max", "Percentage", "Grade"]
ws_data.append(headers)

rows = [
    ["Alex", "1E1", 75],
    ["Bella", "1E2", 45],
    ["Chris", "1E1", 90],
    ["Diana", "1E3", 38],
    ["Ethan", "1E2", 82],
    ["Fiona", "1E1", 59],
    ["Gwen", "1E3", 61],
    ["Harun", "1E2", 70],
    ["Iris", "1E1", 84],
    ["Jamal", "1E2", 49],
    ["Kai", "1E3", 33],
    ["Lina", "1E1", 68],
    ["Mina", "1E2", 88],
    ["Nate", "1E3", 54],
    ["Omar", "1E1", 72],
    ["Pia", "1E2", 96],
    ["Qin", "1E3", 64],
    ["Ravi", "1E1", 80],
    ["Sara", "1E2", 51],
    ["Troy", "1E3", 42],
]
start_row = 2
for i, (name, clazz, marks) in enumerate(rows, start=start_row):
    ws_data[f"A{i}"] = name
    ws_data[f"B{i}"] = clazz
    ws_data[f"C{i}"] = marks
    ws_data[f"D{i}"] = 100  # Max
    ws_data[f"E{i}"] = f"=C{i}/D{i}"  # Percentage
    ws_data[f"F{i}"] = (
        f'=IF(C{i}>=80,"A",IF(C{i}>=70,"B",IF(C{i}>=60,"C",IF(C{i}>=50,"D","U"))))'
    )

# Style header
for col in range(1, 7):
    cell = ws_data.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="F2F2F2")
    cell.border = thin_border

# Percentage format
for i in range(start_row, start_row + len(rows)):
    ws_data[f"E{i}"].number_format = "0%"

# Table
last_row = start_row + len(rows) - 1
table_ref = f"A1:F{last_row}"
add_table(ws_data, table_ref, "tblMarks")

# Column widths, freeze
set_col_widths(ws_data, {"A": 16, "B": 10, "C": 10, "D": 8, "E": 12, "F": 10})
ws_data.freeze_panes = "A2"

# ---------- Conditional Formatting (pre-applied for reference) ----------
marks_range = f"C{start_row}:C{last_row}"

# 1) Marks < 50 : light red fill with dark red text
ws_data.conditional_formatting.add(
    marks_range,
    CellIsRule(
        operator="lessThan",
        formula=["50"],
        stopIfTrue=False,
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        font=Font(color="9C0006", bold=False),
    ),
)

# 2) Marks >= 80 : green fill with dark green text
ws_data.conditional_formatting.add(
    marks_range,
    CellIsRule(
        operator="greaterThanOrEqual",
        formula=["80"],
        stopIfTrue=False,
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        font=Font(color="006100", bold=False),
    ),
)

# 3) Between 40 and 60 : yellow fill
ws_data.conditional_formatting.add(
    marks_range,
    CellIsRule(
        operator="between",
        formula=["40", "60"],
        stopIfTrue=False,
        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    ),
)

# 4) Top 3 marks: formula rule instead of Top10Rule
# Formula is relative to the top-left cell in the applied range.
# =C2>=LARGE($C$2:$C$<last_row>,3)
formula_top3 = f"C{start_row}>=LARGE($C${start_row}:$C${last_row},3)"
rule_top3 = Rule(type="expression", formula=[formula_top3])
rule_top3.dxf = DifferentialStyle(
    fill=PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
)
ws_data.conditional_formatting.add(marks_range, rule_top3)

# 5) Data bar : show relative size (ARGB color string)
data_bar = DataBarRule(
    start_type="num",
    start_value=0,
    end_type="num",
    end_value=100,
    color="FF638EC6",  # must be a quoted string
    showValue=True,
)
ws_data.conditional_formatting.add(marks_range, data_bar)

# ---------- Simple Chart on Data ----------
chart = BarChart()
chart.title = "Marks by Student"
chart.y_axis.title = "Marks"
chart.x_axis.title = "Student"

values = Reference(
    ws_data, min_col=3, min_row=1, max_row=last_row
)  # Marks (include header)
cats = Reference(ws_data, min_col=1, min_row=2, max_row=last_row)  # Names
chart.add_data(values, titles_from_data=True)
chart.set_categories(cats)
chart.height = 10
chart.width = 20
ws_data.add_chart(chart, "H3")

# ---------- Tasks ----------
title(ws_tasks, "Practice Tasks: Conditional Formatting")
tasks = [
    "1) Highlight Marks < 50 with a light red fill and dark red text.",
    "2) Highlight Marks ≥ 80 with a green fill.",
    "3) Highlight Marks between 40 and 60 with a yellow fill.",
    "4) Highlight the Top 3 marks with a blue fill (use a formula rule).",
    "5) Add Data Bars to the Marks column.",
    "6) Bonus: Highlight duplicate Class codes in column B.",
    '7) Bonus: Use a formula rule to highlight Grades = "A" in column F.',
]
for i, t in enumerate(tasks, start=3):
    ws_tasks[f"A{i}"] = t
    ws_tasks[f"A{i}"].alignment = Alignment(wrap_text=True)
set_col_widths(ws_tasks, {"A": 90})

# ---------- Hints ----------
title(ws_hints, "Hints")
marks_hint_range = f"C2:C{last_row}"
hint_rows = [
    (
        "Task 1",
        f"Select {marks_hint_range} → Home → Conditional Formatting → Highlight Cell Rules → Less Than → 50 → pick red fill.",
    ),
    (
        "Task 2",
        f"Select {marks_hint_range} → Highlight Cell Rules → Greater Than or Equal To → 80 → pick green fill.",
    ),
    (
        "Task 3",
        f"Select {marks_hint_range} → Highlight Cell Rules → Between → 40 and 60 → yellow fill.",
    ),
    (
        "Task 4",
        f"Select {marks_hint_range} → New Rule → Use a formula → =C2>=LARGE($C$2:$C${last_row},3) → blue fill.",
    ),
    ("Task 5", f"Select {marks_hint_range} → Data Bars → Gradient Fill (any color)."),
    ("Bonus 6", f"Select B2:B{last_row} → Highlight Cell Rules → Duplicate Values."),
    (
        "Bonus 7",
        f'Select F2:F{last_row} → New Rule → Use a formula → =F2="A" → choose a format.',
    ),
]
ws_hints.append(["Task", "Hint"])
for r in hint_rows:
    ws_hints.append(list(r))
add_table(ws_hints, "A1:B8", "tblHints")
set_col_widths(ws_hints, {"A": 16, "B": 90})
ws_hints.freeze_panes = "A2"

# ---------- Answers ----------
title(ws_answers, "Answer Checks (Helper Columns)")
ws_answers["A3"] = (
    "These formulas evaluate which rows meet each rule on the Data sheet."
)
ws_answers["A3"].alignment = Alignment(wrap_text=True)

# SAFE HEADERS (avoid symbols like < or ≥)
ws_answers.append(
    [
        "Name",
        "Marks",
        "LessThan50",
        "GreaterOrEqual80",
        "Between40_60",
        "Top3",
        "GradeA",
        "ClassDuplicate",
    ]
)
ans_header_row = ws_answers.max_row

# Link formulas back to Data sheet
ans_start = ws_answers.max_row + 1
for i, _ in enumerate(rows, start=2):
    ans_row = ans_start + (i - 2)
    ws_answers[f"A{ans_row}"] = f"=Data!A{i}"
    ws_answers[f"B{ans_row}"] = f"=Data!C{i}"
    ws_answers[f"C{ans_row}"] = f"=IF(Data!C{i}<50,TRUE,FALSE)"
    ws_answers[f"D{ans_row}"] = f"=IF(Data!C{i}>=80,TRUE,FALSE)"
    ws_answers[f"E{ans_row}"] = f"=AND(Data!C{i}>=40,Data!C{i}<=60)"
    ws_answers[f"F{ans_row}"] = (
        f"=IF(Data!C{i}>=LARGE(Data!$C$2:Data!$C${last_row},3),TRUE,FALSE)"
    )
    ws_answers[f"G{ans_row}"] = f'=IF(Data!F{i}="A",TRUE,FALSE)'
    ws_answers[f"H{ans_row}"] = f"=COUNTIF(Data!$B$2:Data!$B${last_row},Data!B{i})>1"

# Build the table starting on the header row
add_table(ws_answers, f"A{ans_header_row}:H{ws_answers.max_row}", "tblAnswers")
set_col_widths(
    ws_answers, {"A": 16, "B": 10, "C": 14, "D": 18, "E": 16, "F": 10, "G": 10, "H": 16}
)
ws_answers.freeze_panes = "A6"

# ---------- Checklist ----------
title(ws_check, "Checklist")
check_items = [
    ("Select a range before adding a rule", ""),
    ("Use Less Than (50) on Marks", ""),
    ("Use Greater Than or Equal (80) on Marks", ""),
    ("Use Between (40,60) on Marks", ""),
    ("Apply Top 3 rule on Marks (formula)", ""),
    ("Add Data Bars to Marks", ""),
    ("(Bonus) Duplicate Values on Class", ""),
    ('(Bonus) Formula rule for Grade = "A"', ""),
]
ws_check.append(["Item", "Done (Y/N)"])
for item, done in check_items:
    ws_check.append([item, done])
add_table(ws_check, f"A1:B{1 + len(check_items) + 1}", "tblChecklist")
set_col_widths(ws_check, {"A": 60, "B": 12})
ws_check.freeze_panes = "A2"

# ---------- Lookup ----------
title(ws_lookup, "Reference: Conditional Formatting Rule Types")
lookup_rows = [
    (
        "Highlight Cell Rules",
        "Greater Than, Less Than, Between, Equal To, Text, Dates, Duplicate",
    ),
    (
        "Top/Bottom Rules",
        "Top N, Bottom N, Above/Below Average (or use LARGE with a formula)",
    ),
    ("Data Bars", "Gradient/Solid bars showing size"),
    ("Color Scales", "2- or 3-color scales"),
    ("Icon Sets", "Arrows, flags, traffic lights etc."),
    ("Use a formula", 'Custom logic like =F2="A"'),
]
ws_lookup.append(["Category", "Examples / Notes"])
for r in lookup_rows:
    ws_lookup.append(list(r))
add_table(ws_lookup, f"A1:B{1 + len(lookup_rows) + 1}", "tblLookup")
set_col_widths(ws_lookup, {"A": 30, "B": 90})
ws_lookup.freeze_panes = "A2"

# ---------- Finishing touches ----------
for ws in [ws_instr, ws_data, ws_tasks, ws_hints, ws_answers, ws_check, ws_lookup]:
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

# Save
wb.save("Conditional_Formatting_Practice.xlsx")
print("Created Conditional_Formatting_Practice.xlsx")
